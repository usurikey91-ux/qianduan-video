import os;

import asyncio
import csv
import hashlib
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
import tempfile
import traceback
import uuid
import urllib.error
import urllib.request
from datetime import datetime
from functools import wraps
from pathlib import Path
from queue import Empty, Queue
from flask_cors import CORS
from itsdangerous import URLSafeTimedSerializer
from myUtils.douyin_benchmark import (
    discover_douyin_benchmark_accounts,
    normalize_douyin_user_url,
    scrape_douyin_benchmark,
)
from flask import Flask, request, jsonify, Response, render_template, send_from_directory
from backend_app.agent.gateway_client import (
    get_hermes_settings as agent_get_hermes_settings,
    hermes_request as agent_hermes_request,
    public_hermes_settings as agent_public_hermes_settings,
)
from backend_app.agent.model_registry import (
    get_agent_models as agent_get_agent_models,
    get_task_agent_model as agent_get_task_agent_model,
    normalize_agent_model as agent_normalize_agent_model,
)
from backend_app.agent.structured_runner import (
    call_universal_ai_structured as agent_call_universal_ai_structured,
    call_hermes_structured as agent_call_hermes_structured,
    load_json_object as agent_load_json_object,
    resolve_executable as agent_resolve_executable,
    run_codex_cli_structured as agent_run_codex_cli_structured,
)
from backend_app.agent.openai_compatible_client import (
    SUPPORTED_AI_PROTOCOLS,
    get_universal_ai_settings,
    public_universal_ai_settings,
)
from backend_app.modules.benchmark.prompts import build_video_analysis_prompt
from backend_app.modules.benchmark import repository as benchmark_repository
from backend_app.modules.benchmark.schemas import video_analysis_schema
from backend_app.modules.opencli_monitor import service as opencli_monitor_service
from backend_app.modules.video_jiexi import client as video_jiexi_client
from backend_app.modules.idea_radar.prompts import build_transcript_radar_prompt as build_idea_radar_prompt
from backend_app.modules.idea_radar import repository as idea_radar_repository
from backend_app.modules.idea_radar.jobs import IdeaRadarJobRegistry
from backend_app.modules.idea_radar.service import get_status as get_idea_radar_status_payload
from backend_app.modules.idea_radar.service import start_pipeline_task
from backend_app.modules.idea_radar.pipeline import run_pipeline as run_idea_radar_pipeline_core
from backend_app.modules.idea_radar.schemas import transcript_radar_schema
from backend_app.modules.idea_radar import media as idea_radar_media
from backend_app.modules.script_generation.prompts import build_identity_script_prompt
from backend_app.modules.script_generation.schemas import identity_script_schema
from backend_app.modules.mcp_server.tools import describe_tools as describe_mcp_tools
from backend_app.modules.mcp_server.tools import create_tool_handlers
from backend_app.modules.own_content_review import repository as own_content_repository
from backend_app.modules.own_content_review import xiaohongshu_repository as xhs_content_repository
from backend_app.modules.own_content_review import connectors as own_content_connectors
from backend_app.modules.own_content_review import platform_connections
from backend_app.modules.own_content_review.service import review_published_content
from backend_app.modules.accounts import repository as account_repository
from backend_app.modules.materials import repository as material_repository
from backend_app.modules.materials import service as material_service
from backend_app.common.database import ensure_core_tables as ensure_base_tables
from backend_app.modules.auth import local_auth
from conf import BASE_DIR


idea_radar_job_registry = IdeaRadarJobRegistry()
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SAU_SECRET_KEY") or secrets.token_hex(32)
AUTH_TOKEN_MAX_AGE = int(os.environ.get("SAU_AUTH_TOKEN_MAX_AGE", str(7 * 24 * 60 * 60)))
AUTH_REQUIRED = os.environ.get("SAU_AUTH_REQUIRED", "0").strip().lower() in {
    "1", "true", "yes", "on"
}
token_serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"], salt="sau-local-auth")

for stream in (sys.stdout, sys.stderr):
    try:
        stream.reconfigure(errors="replace")
    except Exception:
        pass

#允许所有来源跨域访问
CORS(app)

# 限制上传文件大小为160MB
app.config['MAX_CONTENT_LENGTH'] = 160 * 1024 * 1024


def get_db_path():
    return Path(BASE_DIR / "db" / "database.db")


def get_settings_path():
    return Path(BASE_DIR / "settings.json")


def load_runtime_settings():
    settings_path = get_settings_path()
    if not settings_path.exists():
        return {}
    try:
        data = json.loads(settings_path.read_text(encoding="utf-8-sig"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_runtime_settings(settings):
    settings_path = get_settings_path()
    settings_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = settings_path.with_suffix(".json.tmp")
    temporary_path.write_text(
        json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporary_path.replace(settings_path)


DEFAULT_BENCHMARK_MONITORING_RULES = {
    "reference_work_count": 20,
    # OpenCLI Admin still expects the legacy hot/very-hot pair.  The workbench
    # exposes only hot_multiple as the single inclusion threshold; the second
    # value is kept half a step above it solely for connector compatibility.
    "hot_multiple": 5.0,
    "very_hot_multiple": 5.5,
    "interval_hours": 4,
    "inherit_global": True,
}


def normalize_benchmark_monitoring_rules(payload=None, *, inherit_global=True):
    source = payload if isinstance(payload, dict) else {}
    legacy_hot = float(source.get("hot_multiple", source.get("hotMultiple", 5)))
    legacy_very_hot = float(source.get("very_hot_multiple", source.get("veryHotMultiple", 5.5)))
    # Migrate the former default (3x hot / 5x very hot) to the new single 5x
    # inclusion threshold. Other saved custom values keep their lower bound.
    legacy_pair = legacy_hot == 3.0 and (
        legacy_very_hot == 5.0
        or ("very_hot_multiple" not in source and "veryHotMultiple" not in source)
    )
    entry_multiple = 5.0 if legacy_pair else legacy_hot
    rules = {
        "reference_work_count": int(source.get("reference_work_count", source.get("referenceWorkCount", 20))),
        "hot_multiple": entry_multiple,
        "very_hot_multiple": entry_multiple + 0.5,
        "interval_hours": int(source.get("interval_hours", source.get("intervalHours", 4))),
        "inherit_global": bool(source.get("inherit_global", source.get("inheritGlobal", inherit_global))),
    }
    if not 5 <= rules["reference_work_count"] <= 50:
        raise ValueError("参考作品数必须在 5 到 50 条之间")
    if not 1.5 <= rules["hot_multiple"] <= 10:
        raise ValueError("入选倍数必须在 1.5 到 10 之间")
    if rules["interval_hours"] not in {1, 2, 4, 8, 12, 24}:
        raise ValueError("巡检频率只支持 1、2、4、8、12 或 24 小时")
    return rules


def get_benchmark_monitoring_defaults(settings=None):
    settings = settings or load_runtime_settings()
    stored = settings.get("benchmarkMonitoringDefaults")
    try:
        return normalize_benchmark_monitoring_rules(stored, inherit_global=True)
    except (TypeError, ValueError):
        return dict(DEFAULT_BENCHMARK_MONITORING_RULES)


def get_hermes_settings(settings=None):
    return agent_get_hermes_settings(settings or load_runtime_settings())


def get_agent_models(settings=None):
    return agent_get_agent_models(settings or load_runtime_settings())


def public_hermes_settings(settings=None):
    return agent_public_hermes_settings(settings or load_runtime_settings())


def hermes_request(path, method="GET", payload=None, timeout=None, settings=None):
    return agent_hermes_request(path, method=method, payload=payload, timeout=timeout, settings=settings or load_runtime_settings())


def normalize_agent_model(payload, existing_id=None):
    return agent_normalize_agent_model(payload, existing_id)


def get_runtime_setting(*keys, env=None, default=None):
    if env and os.environ.get(env):
        return os.environ.get(env)
    settings = load_runtime_settings()
    for key in keys:
        value = settings.get(key)
        if value:
            return value
    return default


def resolve_executable(command, label):
    return agent_resolve_executable(command, label, get_settings_path())


LOCAL_CODEX_MODEL_ID = "local-codex-default"
UNIVERSAL_AI_MODEL_ID = "universal-ai-default"
LEGACY_OPENAI_COMPATIBLE_MODEL_ID = "openai-compatible-default"


def get_codex_cli_status():
    configured = get_runtime_setting(
        "codexCliPath", "codex_cli_path", env="CODEX_CLI_PATH",
        default="codex.cmd" if os.name == "nt" else "codex",
    )
    try:
        executable = resolve_executable(configured, "Codex CLI")
        version_result = subprocess.run(
            [executable, "--version"], capture_output=True, text=True,
            encoding="utf-8", errors="replace", timeout=15,
        )
        login_result = subprocess.run(
            [executable, "login", "status"], capture_output=True, text=True,
            encoding="utf-8", errors="replace", timeout=20,
        )
        login_text = (login_result.stdout or login_result.stderr or "").strip()
        authenticated = login_result.returncode == 0 and "logged in" in login_text.lower()
        return {
            "available": version_result.returncode == 0,
            "authenticated": authenticated,
            "executable": executable,
            "version": (version_result.stdout or version_result.stderr or "").strip(),
            "loginStatus": login_text,
        }
    except Exception as exc:
        return {
            "available": False, "authenticated": False, "executable": "",
            "version": "", "loginStatus": "", "error": str(exc),
        }


def configure_local_codex_model(settings=None, model_name=None):
    settings = settings or load_runtime_settings()
    status = get_codex_cli_status()
    if not status.get("available"):
        raise RuntimeError(status.get("error") or "未检测到 Codex CLI")
    if not status.get("authenticated"):
        raise RuntimeError("Codex CLI 尚未登录，请先运行 codex login")
    model_name = str(
        model_name
        or get_runtime_setting("codexAnalysisModel", "codex_model", env="CODEX_ANALYSIS_MODEL", default="gpt-5.4-mini")
    ).strip()
    model = normalize_agent_model({
        "id": LOCAL_CODEX_MODEL_ID,
        "name": "本机 Codex",
        "provider": "codex-cli",
        "model": model_name,
        "reasoningEffort": "medium",
        "enabled": True,
    }, existing_id=LOCAL_CODEX_MODEL_ID)
    models = get_agent_models(settings)
    index = next((i for i, item in enumerate(models) if item.get("id") == LOCAL_CODEX_MODEL_ID), None)
    if index is None:
        models.append(model)
    else:
        models[index] = model
    settings["agentModels"] = models
    task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
    if not task_models.get("viralAnalysis"):
        task_models["viralAnalysis"] = LOCAL_CODEX_MODEL_ID
    settings["taskModels"] = task_models
    save_runtime_settings(settings)
    return model


def configure_universal_ai_model(settings=None, model_name=None, provider_name=None):
    settings = settings or load_runtime_settings()
    model_name = str(model_name or "gpt-5.6-sol").strip()
    provider_name = str(provider_name or get_universal_ai_settings(settings)["providerName"]).strip() or "自定义 AI"
    if not model_name:
        raise RuntimeError("模型名称不能为空")
    model = normalize_agent_model({
        "id": UNIVERSAL_AI_MODEL_ID,
        "name": provider_name,
        "provider": "universal-ai",
        "model": model_name,
        "reasoningEffort": "high",
        "enabled": True,
    }, existing_id=UNIVERSAL_AI_MODEL_ID)
    models = [
        item for item in get_agent_models(settings)
        if item.get("id") != LEGACY_OPENAI_COMPATIBLE_MODEL_ID
    ]
    index = next((i for i, item in enumerate(models) if item.get("id") == UNIVERSAL_AI_MODEL_ID), None)
    if index is None:
        models.append(model)
    else:
        models[index] = model
    settings["agentModels"] = models
    task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
    task_models["viralAnalysis"] = UNIVERSAL_AI_MODEL_ID
    settings["taskModels"] = task_models
    save_runtime_settings(settings)
    return model


def ensure_default_agent_model():
    settings = load_runtime_settings()
    if get_agent_models(settings):
        return None
    status = get_codex_cli_status()
    if status.get("available") and status.get("authenticated"):
        return configure_local_codex_model(settings)
    return None


def ensure_local_auth_table():
    local_auth.ensure_admin_table(get_db_path())


def ensure_core_tables():
    ensure_local_auth_table()
    ensure_base_tables(get_db_path())
    ensure_douyin_benchmark_tables()
    ensure_douyin_own_tables()
    ensure_xiaohongshu_own_tables()


def create_auth_token(admin):
    return local_auth.create_token(token_serializer, admin)


def parse_auth_token():
    return local_auth.parse_token(
        token_serializer,
        auth_header=request.headers.get("Authorization", ""),
        query_token=request.args.get("token"),
        max_age=AUTH_TOKEN_MAX_AGE,
    )


def auth_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not AUTH_REQUIRED:
            request.current_admin = {
                "id": "local-user",
                "username": "local",
                "display_name": "本机用户",
            }
            return fn(*args, **kwargs)
        admin = parse_auth_token()
        if not admin:
            return jsonify({"code": 401, "message": "未登录或登录已过期", "data": None}), 401
        request.current_admin = admin
        return fn(*args, **kwargs)
    return wrapper


@app.before_request
def require_local_login():
    if request.method == "OPTIONS":
        return None
    if not AUTH_REQUIRED:
        return None
    public_paths = {
        "/",
        "/favicon.ico",
        "/auth/login",
        "/runtime/identity",
    }
    if request.path in public_paths or request.path.startswith("/assets/"):
        return None
    if request.path.startswith("/auth/"):
        return None
    if request.path.startswith("/api/koubo/"):
        return None
    admin = parse_auth_token()
    if not admin:
        return jsonify({"code": 401, "message": "未登录或登录已过期", "data": None}), 401
    request.current_admin = admin
    return None


@app.route("/runtime/identity", methods=["GET"])
def runtime_identity():
    response = jsonify({
        "code": 200,
        "data": {
            "service": "content-workbench-backend",
            "packaged": os.environ.get("SAU_PACKAGED") == "1",
            "authRequired": AUTH_REQUIRED,
        },
    })
    response.headers["X-SAU-Instance-Token"] = os.environ.get("SAU_INSTANCE_TOKEN", "")
    return response


@app.route("/auth/login", methods=["POST"])
def auth_login():
    ensure_local_auth_table()
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"code": 400, "message": "请输入用户名和密码", "data": None}), 400

    admin = local_auth.get_admin_by_username(get_db_path(), username)
    if not local_auth.verify_admin(admin, password):
        return jsonify({"code": 401, "message": "用户名或密码错误", "data": None}), 401

    user = local_auth.public_user(admin)
    return jsonify({"code": 200, "message": "登录成功", "data": {"token": create_auth_token(admin), "user": user}})


@app.route("/auth/me", methods=["GET"])
@auth_required
def auth_me():
    admin = request.current_admin
    return jsonify({
        "code": 200,
        "message": None,
        "data": {
            "id": admin["id"],
            "username": admin["username"],
            "displayName": admin["display_name"],
        },
    })


@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    return jsonify({"code": 200, "message": "已退出", "data": None})

@app.route("/settings/hermes", methods=["GET", "PUT"])
def hermes_settings_api():
    settings = load_runtime_settings()
    if request.method == "GET":
        return jsonify({"code": 200, "data": public_hermes_settings(settings)})
    payload = request.get_json(silent=True) or {}
    current = get_hermes_settings(settings)
    gateway_url = str(payload.get("gatewayUrl") or "").strip().rstrip("/")
    if not re.match(r"^https?://", gateway_url, re.IGNORECASE):
        return jsonify({"code": 400, "message": "Gateway 地址必须以 http:// 或 https:// 开头"}), 400
    try:
        timeout = max(10, min(int(payload.get("timeout") or current["timeout"]), 1800))
    except (TypeError, ValueError):
        return jsonify({"code": 400, "message": "超时时间必须是数字"}), 400
    api_key = current["apiKey"]
    if payload.get("clearApiKey"):
        api_key = ""
    elif payload.get("apiKey"):
        api_key = str(payload.get("apiKey")).strip()
    settings["hermes"] = {"gatewayUrl": gateway_url, "apiKey": api_key, "timeout": timeout}
    save_runtime_settings(settings)
    return jsonify({"code": 200, "message": "Hermes 配置已保存", "data": public_hermes_settings(settings)})


@app.route("/settings/universal-ai", methods=["GET", "PUT"])
@app.route("/settings/openai-compatible", methods=["GET", "PUT"])
def universal_ai_settings_api():
    settings = load_runtime_settings()
    selected = (settings.get("taskModels") or {}).get("viralAnalysis")
    model = next(
        (item for item in get_agent_models(settings) if item.get("id") in {
            UNIVERSAL_AI_MODEL_ID, LEGACY_OPENAI_COMPATIBLE_MODEL_ID
        }),
        None,
    )
    if request.method == "GET":
        return jsonify({"code": 200, "data": {
            **public_universal_ai_settings(settings),
            "model": (model or {}).get("model") or "gpt-5.6-sol",
            "selectedForViralAnalysis": selected in {
                UNIVERSAL_AI_MODEL_ID, LEGACY_OPENAI_COMPATIBLE_MODEL_ID
            },
        }})

    payload = request.get_json(silent=True) or {}
    current = get_universal_ai_settings(settings)
    provider_name = str(payload.get("providerName") or current["providerName"]).strip() or "自定义 AI"
    protocol = str(payload.get("protocol") or current["protocol"]).strip().lower()
    if protocol not in SUPPORTED_AI_PROTOCOLS:
        return jsonify({"code": 400, "message": "请选择受支持的 AI 接口协议", "data": None}), 400
    base_url = str(payload.get("baseUrl") or "").strip().rstrip("/")
    if not re.match(r"^https?://", base_url, re.IGNORECASE):
        return jsonify({"code": 400, "message": "AI 接口地址必须以 http:// 或 https:// 开头", "data": None}), 400
    try:
        timeout = max(10, min(int(payload.get("timeout") or current["timeout"]), 1800))
    except (TypeError, ValueError):
        return jsonify({"code": 400, "message": "超时时间必须是数字", "data": None}), 400
    api_key = current["apiKey"]
    if payload.get("clearApiKey"):
        api_key = ""
    elif payload.get("apiKey"):
        api_key = str(payload.get("apiKey")).strip()
    settings["universalAI"] = {
        "providerName": provider_name, "protocol": protocol,
        "baseUrl": base_url, "apiKey": api_key, "timeout": timeout,
    }
    settings.pop("openaiCompatible", None)
    model = configure_universal_ai_model(settings, payload.get("model"), provider_name)
    settings = load_runtime_settings()
    return jsonify({"code": 200, "message": "通用 AI 模型已保存并设为爆款拆解模型", "data": {
        **public_universal_ai_settings(settings),
        "model": model.get("model"),
        "selectedForViralAnalysis": True,
    }})


@app.route("/settings/universal-ai/test", methods=["POST"])
@app.route("/settings/openai-compatible/test", methods=["POST"])
def test_universal_ai_settings():
    settings = load_runtime_settings()
    model = next(
        (item for item in get_agent_models(settings) if item.get("id") in {
            UNIVERSAL_AI_MODEL_ID, LEGACY_OPENAI_COMPATIBLE_MODEL_ID
        }),
        None,
    )
    if not model:
        return jsonify({"code": 400, "message": "请先保存通用 AI 模型配置", "data": None}), 400
    started = time.monotonic()
    try:
        result = agent_call_universal_ai_structured(
            "只返回一个 JSON 对象，字段 result 的值必须是 OK。",
            {"type": "object", "additionalProperties": False, "required": ["result"],
             "properties": {"result": {"type": "string"}}},
            settings=settings, model_config=model, timeout=120, log=backend_log,
        )
        return jsonify({"code": 200, "message": "通用 AI 模型连接正常", "data": {
            "result": result, "elapsedMs": int((time.monotonic() - started) * 1000),
            "provider": model.get("provider"), "model": model.get("model"),
        }})
    except Exception as exc:
        return jsonify({"code": 502, "message": str(exc), "data": None}), 502


@app.route("/settings/integrations", methods=["GET", "PUT"])
def integrations_settings_api():
    settings = load_runtime_settings()
    if request.method == "GET":
        return jsonify({"code": 200, "message": None, "data": {
            "opencliAdminBaseUrl": opencli_monitor_service.get_base_url(settings),
            "opencliAdminApiTokenConfigured": bool(opencli_monitor_service.get_api_token(settings)),
            "videoJiexiBaseUrl": video_jiexi_client.base_url(settings),
            "videoJiexiApiTokenConfigured": bool(video_jiexi_client.api_token(settings)),
            "videoJiexiDownloadDir": str(video_jiexi_client.download_root(settings) or ""),
            "factsOnlyMode": settings.get("factsOnlyMode") is True,
        }})

    payload = request.get_json(silent=True) or {}
    opencli_url = str(payload.get("opencliAdminBaseUrl") or "").strip().rstrip("/")
    video_url = str(payload.get("videoJiexiBaseUrl") or "").strip().rstrip("/")
    for label, value in (("OpenCLI Admin", opencli_url), ("video-jiexi", video_url)):
        if value and not re.match(r"^https?://", value, re.IGNORECASE):
            return jsonify({"code": 400, "message": f"{label} 地址必须以 http:// 或 https:// 开头", "data": None}), 400

    settings["opencliAdminBaseUrl"] = opencli_url
    settings["videoJiexiBaseUrl"] = video_url
    if payload.get("clearOpencliAdminApiToken"):
        settings.pop("opencliAdminApiToken", None)
    elif payload.get("opencliAdminApiToken"):
        settings["opencliAdminApiToken"] = str(payload.get("opencliAdminApiToken")).strip()
    if payload.get("clearVideoJiexiApiToken"):
        settings.pop("videoJiexiApiToken", None)
    elif payload.get("videoJiexiApiToken"):
        settings["videoJiexiApiToken"] = str(payload.get("videoJiexiApiToken")).strip()
    download_dir = str(payload.get("videoJiexiDownloadDir") or "").strip()
    if download_dir:
        settings["videoJiexiDownloadDir"] = download_dir
    else:
        settings.pop("videoJiexiDownloadDir", None)
    save_runtime_settings(settings)
    return jsonify({"code": 200, "message": "集成服务配置已保存", "data": {
        "opencliAdminBaseUrl": opencli_monitor_service.get_base_url(settings),
        "opencliAdminApiTokenConfigured": bool(opencli_monitor_service.get_api_token(settings)),
        "videoJiexiBaseUrl": video_jiexi_client.base_url(settings),
        "videoJiexiApiTokenConfigured": bool(video_jiexi_client.api_token(settings)),
        "videoJiexiDownloadDir": str(video_jiexi_client.download_root(settings) or ""),
        "factsOnlyMode": settings.get("factsOnlyMode") is True,
    }})


@app.route("/settings/hermes/test", methods=["POST"])
def test_hermes_settings():
    try:
        health = hermes_request("/health", timeout=15)
        capabilities = None
        try:
            capabilities = hermes_request("/v1/capabilities", timeout=20)
        except Exception as exc:
            backend_log(f"Hermes capabilities unavailable, using legacy mode: {exc}")
        return jsonify({"code": 200, "message": "Hermes Gateway 连接正常", "data": {
            "health": health, "capabilities": capabilities,
            "legacyMode": capabilities is None,
        }})
    except Exception as exc:
        return jsonify({"code": 502, "message": str(exc), "data": None}), 502


@app.route("/settings/hermes/models", methods=["GET"])
def discover_hermes_models():
    try:
        refresh = request.args.get("refresh") in {"1", "true", "True"}
        path = "/api/model/options?refresh=1" if refresh else "/api/model/options"
        try:
            options = hermes_request(path, timeout=60 if refresh else 30)
        except Exception as exc:
            backend_log(f"Hermes rich model catalog unavailable, using /v1/models: {exc}")
            legacy = hermes_request("/v1/models", timeout=30)
            rows = legacy.get("data") if isinstance(legacy, dict) else []
            options = {
                "legacyMode": True,
                "providers": [{
                    "provider": "gateway-default",
                    "models": [
                        {"id": item.get("id"), "name": item.get("id")}
                        for item in (rows or []) if isinstance(item, dict) and item.get("id")
                    ],
                }],
            }
        return jsonify({"code": 200, "data": options})
    except Exception as exc:
        return jsonify({"code": 502, "message": str(exc), "data": None}), 502


@app.route("/settings/codex-cli", methods=["GET", "POST"])
def codex_cli_settings_api():
    status = get_codex_cli_status()
    if request.method == "GET":
        settings = load_runtime_settings()
        selected = (settings.get("taskModels") or {}).get("viralAnalysis")
        model = next(
            (item for item in get_agent_models(settings) if item.get("id") == LOCAL_CODEX_MODEL_ID),
            None,
        )
        return jsonify({"code": 200, "data": {
            **status,
            "configured": bool(model),
            "selectedForViralAnalysis": selected == LOCAL_CODEX_MODEL_ID,
            "model": model,
        }})
    try:
        payload = request.get_json(silent=True) or {}
        model = configure_local_codex_model(model_name=payload.get("model"))
        settings = load_runtime_settings()
        task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
        task_models["viralAnalysis"] = LOCAL_CODEX_MODEL_ID
        settings["taskModels"] = task_models
        save_runtime_settings(settings)
        return jsonify({"code": 200, "message": "本机 Codex 已设为爆款拆解模型", "data": {
            **get_codex_cli_status(), "configured": True,
            "selectedForViralAnalysis": True, "model": model,
        }})
    except Exception as exc:
        return jsonify({"code": 400, "message": str(exc), "data": None}), 400


@app.route("/settings/agent-models", methods=["GET", "POST"])
def agent_models_api():
    ensure_default_agent_model()
    settings = load_runtime_settings()
    if request.method == "GET":
        task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
        return jsonify({"code": 200, "data": {"models": get_agent_models(settings), "taskModels": task_models}})
    try:
        model = normalize_agent_model(request.get_json(silent=True) or {})
        models = get_agent_models(settings)
        models.append(model)
        settings["agentModels"] = models
        save_runtime_settings(settings)
        return jsonify({"code": 200, "message": "Agent 模型已添加", "data": model})
    except ValueError as exc:
        return jsonify({"code": 400, "message": str(exc), "data": None}), 400


@app.route("/settings/agent-models/<model_id>", methods=["PUT", "DELETE"])
def agent_model_api(model_id):
    settings = load_runtime_settings()
    models = get_agent_models(settings)
    index = next((i for i, item in enumerate(models) if item.get("id") == model_id), None)
    if index is None:
        return jsonify({"code": 404, "message": "Agent 模型不存在", "data": None}), 404
    if request.method == "DELETE":
        models.pop(index)
        task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
        settings["taskModels"] = {key: value for key, value in task_models.items() if value != model_id}
        settings["agentModels"] = models
        save_runtime_settings(settings)
        return jsonify({"code": 200, "message": "Agent 模型已删除", "data": None})
    try:
        model = normalize_agent_model(request.get_json(silent=True) or {}, existing_id=model_id)
        models[index] = model
        settings["agentModels"] = models
        save_runtime_settings(settings)
        return jsonify({"code": 200, "message": "Agent 模型已更新", "data": model})
    except ValueError as exc:
        return jsonify({"code": 400, "message": str(exc), "data": None}), 400


@app.route("/settings/task-models", methods=["PUT"])
def task_models_api():
    payload = request.get_json(silent=True) or {}
    viral_model_id = str(payload.get("viralAnalysis") or "").strip()
    settings = load_runtime_settings()
    if viral_model_id and not any(
        item.get("id") == viral_model_id and item.get("enabled", True)
        for item in get_agent_models(settings)
    ):
        return jsonify({"code": 400, "message": "所选 Agent 模型不存在或已停用", "data": None}), 400
    task_models = settings.get("taskModels") if isinstance(settings.get("taskModels"), dict) else {}
    task_models["viralAnalysis"] = viral_model_id
    settings["taskModels"] = task_models
    save_runtime_settings(settings)
    return jsonify({"code": 200, "message": "爆款拆解模型已保存", "data": task_models})


@app.route("/settings/agent-models/<model_id>/test", methods=["POST"])
def test_agent_model(model_id):
    settings = load_runtime_settings()
    model = next((item for item in get_agent_models(settings) if item.get("id") == model_id), None)
    if not model:
        return jsonify({"code": 404, "message": "Agent 模型不存在", "data": None}), 404
    started = time.monotonic()
    try:
        result = run_agent_structured(
            "只返回一个 JSON 对象，字段 result 的值必须是 OK。",
            {"type": "object", "additionalProperties": False, "required": ["result"],
             "properties": {"result": {"type": "string"}}},
            model_config=model, timeout=120,
        )
        return jsonify({"code": 200, "message": "模型测试成功", "data": {
            "result": result, "elapsedMs": int((time.monotonic() - started) * 1000),
            "provider": model.get("provider"), "model": model.get("model"),
        }})
    except Exception as exc:
        return jsonify({"code": 502, "message": str(exc), "data": None}), 502


def safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        text = " ".join(str(arg) for arg in args)
        encoding = getattr(sys.stdout, "encoding", None) or "utf-8"
        print(text.encode(encoding, errors="replace").decode(encoding, errors="replace"), **kwargs)


def backend_log(message):
    try:
        log_dir = Path(BASE_DIR / "logs")
        log_dir.mkdir(parents=True, exist_ok=True)
        with open(log_dir / "backend.log", "a", encoding="utf-8") as file:
            file.write(f"[{datetime.now().isoformat()}] {message}\n")
    except Exception:
        pass

def ensure_douyin_benchmark_tables():
    benchmark_repository.ensure_tables(get_db_path())


def ensure_douyin_own_tables():
    own_content_repository.ensure_tables(get_db_path())


def ensure_xiaohongshu_own_tables():
    xhs_content_repository.ensure_tables(get_db_path())


def latest_douyin_cookie_file():
    return account_repository.latest_douyin_cookie_file(get_db_path())


OWN_DOUYIN_FIELD_ALIASES = {
    "title": ["作品名称", "笔记名称", "标题", "title", "作品标题", "笔记标题"],
    "video_url": ["作品链接", "笔记链接", "视频链接", "video_url", "url", "链接"],
    "published_at": ["发布时间", "published_at", "发布时间 "],
    "content_format": ["体裁", "内容体裁", "content_format"],
    "visibility_status": ["审核状态", "状态", "visibility_status"],
    "exposure_count": ["曝光量", "曝光次数", "exposure_count"],
    "play_count": ["播放量", "播放次数", "play_count"],
    "completion_rate": ["完播率", "completion_rate"],
    "five_sec_completion_rate": ["5s完播率", "5秒完播率", "five_sec_completion_rate"],
    "cover_click_rate": ["封面点击率", "cover_click_rate"],
    "two_sec_bounce_rate": ["2s跳出率", "2秒跳出率", "two_sec_bounce_rate"],
    "avg_play_duration": ["平均播放时长", "avg_play_duration"],
    "like_count": ["点赞量", "点赞数", "点赞", "like_count"],
    "share_count": ["分享量", "分享数", "分享", "share_count"],
    "comment_count": ["评论量", "评论数", "评论", "comment_count"],
    "collect_count": ["收藏量", "收藏数", "收藏", "collect_count"],
    "profile_visit_count": ["主页访问量", "profile_visit_count"],
    "follower_delta": ["粉丝增量", "涨粉数", "follower_delta"],
    "transcript": ["作品文案", "文案", "口播文案", "transcript"],
    "notes": ["备注", "notes"],
}

OWN_DOUYIN_INTEGER_FIELDS = {
    "exposure_count", "play_count", "like_count", "share_count", "comment_count", "collect_count",
    "profile_visit_count", "follower_delta",
}
OWN_DOUYIN_FLOAT_FIELDS = {
    "completion_rate", "five_sec_completion_rate", "cover_click_rate",
    "two_sec_bounce_rate", "avg_play_duration",
}


def normalize_import_header(value):
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def build_own_douyin_field_map(headers):
    normalized_headers = {normalize_import_header(header): header for header in headers}
    mapping = {}
    for target, aliases in OWN_DOUYIN_FIELD_ALIASES.items():
        for alias in aliases:
            header = normalized_headers.get(normalize_import_header(alias))
            if header is not None:
                mapping[target] = header
                break
    return mapping


def clean_import_value(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"-", "—", "nan", "NaN", "None"}:
        return None
    return text


def parse_import_int(value):
    text = clean_import_value(value)
    if text is None:
        return None
    text = text.replace(",", "")
    try:
        return int(float(text))
    except Exception:
        return None


def parse_import_float(value):
    text = clean_import_value(value)
    if text is None:
        return None
    text = text.replace(",", "")
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100
        return float(text)
    except Exception:
        return None


def normalize_import_row(raw_row, field_map):
    item = {}
    for field, header in field_map.items():
        value = raw_row.get(header)
        if field in OWN_DOUYIN_INTEGER_FIELDS:
            item[field] = parse_import_int(value)
        elif field in OWN_DOUYIN_FLOAT_FIELDS:
            item[field] = parse_import_float(value)
        else:
            item[field] = clean_import_value(value)
    item["title"] = item.get("title") or ""
    return item


def own_douyin_source_key(item):
    url = clean_import_value(item.get("video_url"))
    if url:
        return f"url:{url}"
    title = clean_import_value(item.get("title")) or ""
    published_at = clean_import_value(item.get("published_at")) or ""
    digest = hashlib.sha1(f"{title}|{published_at}".encode("utf-8")).hexdigest()
    return f"title_time:{digest}"


def read_own_douyin_import_file(file_storage):
    filename = (file_storage.filename or "").lower()
    payload = file_storage.read()
    rows = []
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，请先安装 requirements.txt") from exc
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
        for row in values[1:]:
            raw = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
            if any(clean_import_value(value) for value in raw.values()):
                rows.append(raw)
        return headers, rows
    if filename.endswith(".csv"):
        text = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = payload.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        rows = [row for row in reader if any(clean_import_value(value) for value in row.values())]
        return headers, rows
    raise ValueError("仅支持 .xlsx 或 .csv 文件")


def parse_own_douyin_import(file_storage):
    headers, raw_rows = read_own_douyin_import_file(file_storage)
    field_map = build_own_douyin_field_map(headers)
    if "title" not in field_map:
        raise ValueError("导入文件缺少作品名称/标题字段")
    normalized_rows = [normalize_import_row(row, field_map) for row in raw_rows]
    valid_rows = [row for row in normalized_rows if row.get("title")]
    return {
        "headers": headers,
        "field_map": field_map,
        "rows": valid_rows,
        "raw_count": len(raw_rows),
        "valid_count": len(valid_rows),
    }


def upsert_own_douyin_account(name="我的账号"):
    account_name = clean_import_value(name) or "我的账号"
    return own_content_repository.upsert_account(get_db_path(), account_name)


def save_own_douyin_import(rows, account_name="我的账号"):
    account_name = clean_import_value(account_name) or "我的账号"
    return own_content_repository.save_import(
        get_db_path(),
        rows,
        account_name,
        own_douyin_source_key,
    )


def list_own_douyin_videos(limit=100):
    return own_content_repository.list_videos(get_db_path(), limit)


def get_own_douyin_overview():
    return own_content_repository.get_latest_account_snapshot(get_db_path())


def save_own_xiaohongshu_import(rows, account_name="我的小红书账号"):
    account_name = clean_import_value(account_name) or "我的小红书账号"
    return xhs_content_repository.save_import(
        get_db_path(), rows, account_name, own_douyin_source_key,
    )


def list_own_xiaohongshu_videos(limit=100, account_id=None):
    return xhs_content_repository.list_videos(get_db_path(), limit, account_id=account_id)


def get_own_xiaohongshu_overview(account_id=None):
    return xhs_content_repository.get_latest_account_snapshot(get_db_path(), account_id=account_id)


def get_douyin_benchmark_video_urls(account_id):
    return benchmark_repository.list_video_urls(get_db_path(), account_id)


def upsert_douyin_benchmark(homepage_url):
    return benchmark_repository.upsert_account(
        get_db_path(), homepage_url, normalize_douyin_user_url
    )


def delete_douyin_benchmark(account_id):
    pending = benchmark_repository.delete_account_cascade(get_db_path(), account_id)
    if not pending:
        return None
    account = pending["account"]
    active_video_ids = idea_radar_job_registry.cancel_many(pending["video_ids"])
    cancelled_job_count = len(active_video_ids)
    deleted = pending["deleted"]

    return {
        "id": account_id,
        "nickname": account["nickname"],
        "homepage_url": account["homepage_url"],
        "deleted_videos": deleted["videos"],
        "deleted_analyses": deleted["analysis"],
        "deleted_transcripts": deleted["transcripts"],
        "cancelled_jobs": cancelled_job_count,
    }


def save_douyin_benchmark_sync(account_id, data):
    return benchmark_repository.save_sync(get_db_path(), account_id, data)


def save_douyin_benchmark_error(account_id, error_message):
    benchmark_repository.save_error(get_db_path(), account_id, error_message)


def sync_one_douyin_benchmark(homepage_url, cookie_file, max_videos=20):
    homepage_url = normalize_douyin_user_url(homepage_url)
    if not homepage_url:
        raise ValueError("无效的抖音用户主页链接")
    account_id = upsert_douyin_benchmark(homepage_url)
    existing_urls = get_douyin_benchmark_video_urls(account_id)
    sync_data = asyncio.run(scrape_douyin_benchmark(
        homepage_url,
        cookie_file=cookie_file,
        max_videos=max_videos,
        existing_video_urls=existing_urls
    ))
    stats = save_douyin_benchmark_sync(account_id, sync_data)
    return {"id": account_id, "sync": stats, "account": sync_data}


def parse_keywords_payload(value):
    if isinstance(value, list):
        keywords = value
    else:
        keywords = re.split(r"[,，\n\r]+", str(value or ""))
    return [item.strip() for item in keywords if item and item.strip()]


def split_title_parts(title):
    text = re.sub(r"\s+", " ", title or "").strip()
    if not text:
        return []
    parts = re.split(r"[，,。！？!?；;\n\r]+", text)
    return [part.strip() for part in parts if part.strip()]


def is_count_like_text(value):
    text = re.sub(r"\s+", " ", value or "").strip()
    return bool(re.fullmatch(r"[\d.,]+\s*[万wW]?", text))


def pick_analysis_source_text(video):
    raw_data = {}
    try:
        raw_data = json.loads(video.get("raw_data") or "{}")
    except Exception:
        raw_data = {}
    candidates = [
        video.get("title"),
        raw_data.get("title"),
        raw_data.get("raw_text"),
        raw_data.get("text"),
    ]
    for value in candidates:
        text = re.sub(r"\s+", " ", value or "").strip()
        if text and not is_count_like_text(text) and not text.startswith("热门"):
            return text
    return "该作品暂未同步到标题/文案"


def is_douyin_analysis_stale(analysis, video):
    if not analysis:
        return False
    parsed = parse_douyin_video_analysis(analysis)
    hook = parsed.get("hook") or ""
    summary = parsed.get("summary") or ""
    if is_count_like_text(hook):
        return True
    if re.search(r"围绕「[\d.,]+\s*[万wW]?」", summary):
        return True
    source_text = pick_analysis_source_text(video)
    return source_text != "该作品暂未同步到标题/文案" and hook and hook not in source_text


def pick_keywords(text):
    candidates = [
        "AI", "智能体", "Agent", "Coze", "Obsidian", "Hermes", "Claude", "抖音", "视频号",
        "副业", "自媒体", "知识库", "工作流", "私域", "变现", "获客", "剪辑", "口播",
        "爆款", "账号", "流量", "工具", "教程", "开源", "成本", "模型", "算力", "提示词"
    ]
    found = []
    lower_text = (text or "").lower()
    for word in candidates:
        if word.lower() in lower_text and word not in found:
            found.append(word)
    hashtags = re.findall(r"#([^#\s，。！？!?,;；]+)", text or "")
    for tag in hashtags:
        if tag and tag not in found:
            found.append(tag)
    return found[:8]


def generate_douyin_video_analysis_fallback(video):
    source_text = pick_analysis_source_text(video)
    parts = split_title_parts(source_text)
    keywords = pick_keywords(source_text)
    hook = parts[0] if parts else source_text[:40]
    core = parts[1] if len(parts) > 1 else f"围绕「{hook}」展开观点，用具体问题吸引目标用户继续看。"
    pain_points = [
        "观众可能正在寻找更高效的方法或工具",
        "观众需要把抽象概念变成可执行步骤",
    ]
    if any(word in source_text for word in ["成本", "价格", "便宜", "贵"]):
        pain_points.insert(0, "用户关心成本下降后自己能抓住什么机会")
    if any(word in source_text for word in ["学", "教程", "入门", "路径"]):
        pain_points.insert(0, "用户想知道从哪里开始学、按什么路径学")

    viral_points = [
        "选题有明确对象和具体问题，适合做开头钩子",
        "可以拆成短句字幕和重点弹窗，降低理解成本",
    ]
    if keywords:
        viral_points.append(f"关键词可前置：{' / '.join(keywords[:4])}")

    reusable_points = [
        "用“现象 -> 原因 -> 普通人行动”的结构复刻",
        "保留一个具体例子，再给出三条可执行路径",
        "结尾用清单式行动建议，引导收藏或评论",
    ]
    script_suggestions = [
        f"开头：最近我看到一个信号，{hook}",
        "中段：这件事真正影响普通人的地方，不是新闻本身，而是成本和门槛变化",
        "结尾：如果你也想跟上，可以先从一个工具、一个教程、一个项目开始练",
    ]

    analysis = {
        "basis": "metadata",
        "keywords": keywords,
        "structure": [
            {"label": "开头钩子", "content": hook},
            {"label": "核心观点", "content": core},
            {"label": "论证/案例", "content": "用行业变化、工具案例或个人体验证明观点。"},
            {"label": "行动引导", "content": "给出可执行路径，让用户知道下一步做什么。"},
        ],
        "pain_points": pain_points[:4],
        "viral_points": viral_points[:4],
        "reusable_points": reusable_points,
        "script_suggestions": script_suggestions,
    }

    return {
        "analysis_type": "metadata",
        "summary": f"这是一个围绕「{hook}」展开的知识分享选题，适合复刻为观点口播或教程切片。",
        "hook": hook,
        "core_viewpoint": core,
        "pain_points": pain_points[:4],
        "viral_points": viral_points[:4],
        "reusable_points": reusable_points,
        "script_suggestions": script_suggestions,
        "raw_analysis": analysis,
    }


def build_douyin_analysis_prompt(video):
    raw_data = {}
    try:
        raw_data = json.loads(video.get("raw_data") or "{}")
    except Exception:
        raw_data = {}
    return build_video_analysis_prompt(video, pick_analysis_source_text(video), raw_data)


def get_codex_analysis_schema():
    return video_analysis_schema()


def load_json_object(text):
    return agent_load_json_object(text)


def normalize_text_list(value, limit=5):
    if isinstance(value, list):
        items = value
    elif isinstance(value, str) and value.strip():
        items = re.split(r"[\n；;]+", value)
    else:
        items = []
    result = []
    for item in items:
        text = re.sub(r"\s+", " ", str(item)).strip()
        if text and text not in result:
            result.append(text)
    return result[:limit]


def normalize_codex_video_analysis(video, codex_data):
    source_text = pick_analysis_source_text(video)
    hook = re.sub(r"\s+", " ", str(codex_data.get("hook") or "")).strip()
    if not hook or is_count_like_text(hook):
        parts = split_title_parts(source_text)
        hook = parts[0] if parts else source_text[:40]

    raw_analysis = {
        "basis": "codex_cli",
        "keywords": normalize_text_list(codex_data.get("keywords"), 8),
        "structure": codex_data.get("structure") or [],
        "source_text": source_text,
    }
    return {
        "analysis_type": "codex_cli",
        "summary": re.sub(r"\s+", " ", str(codex_data.get("summary") or "")).strip(),
        "hook": hook,
        "core_viewpoint": re.sub(r"\s+", " ", str(codex_data.get("core_viewpoint") or "")).strip(),
        "pain_points": normalize_text_list(codex_data.get("pain_points"), 5),
        "viral_points": normalize_text_list(codex_data.get("viral_points"), 5),
        "reusable_points": normalize_text_list(codex_data.get("reusable_points"), 5),
        "script_suggestions": normalize_text_list(codex_data.get("script_suggestions"), 5),
        "raw_analysis": raw_analysis,
    }


def generate_douyin_video_analysis_by_codex(video):
    prompt = build_douyin_analysis_prompt(video)
    timeout = int(get_runtime_setting(
        "codexAnalysisTimeout", "codex_timeout", env="CODEX_ANALYSIS_TIMEOUT_SECONDS", default="180"
    ))
    codex_cmd = resolve_executable(
        get_runtime_setting("codexCliPath", "codex_cli_path", env="CODEX_CLI_PATH", default="codex"),
        "Codex CLI",
    )
    codex_model = get_runtime_setting(
        "codexAnalysisModel", "codex_model", env="CODEX_ANALYSIS_MODEL", default="gpt-5.4-mini"
    )

    with tempfile.TemporaryDirectory(prefix="douyin-codex-analysis-") as tmp_dir:
        schema_path = Path(tmp_dir) / "schema.json"
        output_path = Path(tmp_dir) / "result.json"
        schema_path.write_text(json.dumps(get_codex_analysis_schema(), ensure_ascii=False), encoding="utf-8")

        command = [
            codex_cmd,
            "exec",
            "--model",
            codex_model,
            "--sandbox",
            "read-only",
            "--ephemeral",
            "--ignore-rules",
            "--cd",
            str(BASE_DIR),
            "--output-schema",
            str(schema_path),
            "--output-last-message",
            str(output_path),
            "-",
        ]
        completed = subprocess.run(
            command,
            input=prompt,
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
        if completed.returncode != 0:
            raise RuntimeError(f"Codex CLI failed: {completed.stderr[-2000:] or completed.stdout[-2000:]}")

        output_text = output_path.read_text(encoding="utf-8") if output_path.exists() else completed.stdout
        codex_data = load_json_object(output_text)
        return normalize_codex_video_analysis(video, codex_data)


def generate_douyin_video_analysis(video):
    try:
        return generate_douyin_video_analysis_by_codex(video)
    except Exception as e:
        analysis = generate_douyin_video_analysis_fallback(video)
        analysis["analysis_type"] = "metadata_fallback"
        raw_analysis = analysis.get("raw_analysis") or {}
        raw_analysis["codex_error"] = str(e)
        analysis["raw_analysis"] = raw_analysis
        return analysis


def save_douyin_video_analysis(video_id, analysis):
    return benchmark_repository.save_analysis(get_db_path(), video_id, analysis)


def parse_douyin_video_analysis(row):
    if not row:
        return None
    item = dict(row)
    for key in ("pain_points", "viral_points", "reusable_points", "script_suggestions"):
        try:
            item[key] = json.loads(item.get(key) or "[]")
        except Exception:
            item[key] = []
    try:
        item["raw_analysis"] = json.loads(item.get("raw_analysis") or "{}")
    except Exception:
        item["raw_analysis"] = {}
    return item


def parse_idea_radar_transcript(row):
    return idea_radar_repository.parse_transcript(row)


def get_idea_radar_transcript(video_id):
    return idea_radar_repository.get_transcript(get_db_path(), video_id)


def update_idea_radar_transcript(video_id, **values):
    idea_radar_repository.update_transcript(get_db_path(), video_id, **values)


def update_idea_radar_progress(video_id, stage, percent, message, status="processing", **values):
    idea_radar_job_registry.ensure_not_cancelled(video_id)
    current = get_idea_radar_transcript(video_id) or {}
    logs = list(current.get("progress_log") or [])
    percent = max(0, min(int(round(percent or 0)), 100))
    last_entry = logs[-1] if logs else {}
    if last_entry.get("message") != message or last_entry.get("percent") != percent:
        logs.append({
            "time": datetime.now().strftime("%H:%M:%S"),
            "stage": stage,
            "percent": percent,
            "message": message,
        })
        logs = logs[-60:]
    payload = {
        "status": status,
        "stage": stage,
        "progress_percent": percent,
        "progress_message": message,
        "progress_log": logs,
        **values,
    }
    update_idea_radar_transcript(video_id, **payload)
    print(f"[观点雷达 {video_id}] {percent}% {message}", flush=True)


def write_netscape_cookie_file(storage_state_path, output_path):
    return idea_radar_media.write_netscape_cookie_file(storage_state_path, output_path)


def run_streaming_command(command, timeout, on_line=None, idle_timeout=None):
    backend_log(f"run command: {' '.join(str(item) for item in command)}")
    process = subprocess.Popen(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
    )
    line_queue = Queue()

    def read_output():
        try:
            for line in process.stdout or []:
                line_queue.put(line.rstrip())
        finally:
            line_queue.put(None)

    reader = threading.Thread(target=read_output, daemon=True)
    reader.start()
    output = []
    deadline = time.time() + timeout
    last_output_at = time.time()
    stream_closed = False
    while not stream_closed or process.poll() is None:
        if time.time() > deadline:
            process.kill()
            raise subprocess.TimeoutExpired(command, timeout)
        if idle_timeout and time.time() - last_output_at > idle_timeout:
            process.kill()
            raise TimeoutError(f"command produced no output for {idle_timeout} seconds")
        try:
            line = line_queue.get(timeout=0.25)
        except Empty:
            continue
        if line is None:
            stream_closed = True
            continue
        output.append(line)
        last_output_at = time.time()
        backend_log(f"command output: {line}")
        if on_line:
            try:
                on_line(line)
            except Exception:
                process.kill()
                process.wait()
                raise
    return process.wait(), output


def download_douyin_video_with_ytdlp(video_url, output_dir, cookie_file, progress_callback=None):
    output_template = str(Path(output_dir) / "source.%(ext)s")
    command = [
        sys.executable, "-m", "yt_dlp", "--ignore-config", "--no-playlist", "--no-warnings",
        "--socket-timeout", "20", "--retries", "2", "--fragment-retries", "2",
        "--newline", "--progress-template",
        "download:%(progress._percent_str)s|%(progress._speed_str)s|%(progress._eta_str)s",
        "--merge-output-format", "mp4", "-o", output_template,
    ]
    ffmpeg_dir = os.environ.get("SAU_FFMPEG_DIR")
    if ffmpeg_dir and Path(ffmpeg_dir).exists():
        command.extend(["--ffmpeg-location", ffmpeg_dir])
    if cookie_file:
        command.extend(["--cookies", str(cookie_file)])
    command.append(video_url)
    def handle_line(line):
        match = re.search(r"download:\s*([\d.]+)%\|([^|]*)\|([^|]*)", line)
        if not match or not progress_callback:
            return
        percent = float(match.group(1))
        speed = match.group(2).strip()
        eta = match.group(3).strip()
        details = [f"下载 {percent:.1f}%"]
        if speed and speed != "N/A":
            details.append(speed)
        if eta and eta != "N/A":
            details.append(f"剩余 {eta}")
        progress_callback(percent, " · ".join(details))

    backend_log(f"yt-dlp ffmpeg dir: {ffmpeg_dir or '<none>'}")
    returncode, output = run_streaming_command(command, timeout=180, idle_timeout=60, on_line=handle_line)
    candidates = sorted(Path(output_dir).glob("source.*"))
    if returncode != 0 or not candidates:
        detail = "\n".join(output[-30:]) or "yt-dlp 下载失败"
        raise RuntimeError(detail)
    if progress_callback:
        progress_callback(100, "视频下载完成")
    return candidates[0]


def download_douyin_video_with_ytdlp(video_url, output_dir, cookie_file, progress_callback=None):
    return idea_radar_media.download_with_ytdlp(
        video_url,
        output_dir,
        cookie_file,
        progress_callback=progress_callback,
        log=backend_log,
    )


async def download_douyin_video_with_playwright(
    video_url, output_path, storage_state_path=None, progress_callback=None
):
    return await idea_radar_media.download_with_playwright(
        video_url,
        output_path,
        storage_state_path=storage_state_path,
        progress_callback=progress_callback,
    )


def download_douyin_video(video_url, work_dir, progress_callback=None):
    return idea_radar_media.download_douyin_video(
        video_url,
        work_dir,
        base_dir=BASE_DIR,
        latest_cookie_file=latest_douyin_cookie_file,
        progress_callback=progress_callback,
        log=backend_log,
    )


def transcribe_idea_radar_media(media_path, progress_callback=None):
    model = os.environ.get("IDEA_RADAR_WHISPER_MODEL", "base")
    script_path = Path(BASE_DIR / "scripts" / "whisper_asr.py")
    command = [
        sys.executable, str(script_path), "--input", str(media_path),
        "--model", model, "--language", "zh", "--device", "cpu", "--compute-type", "int8",
    ]
    result = {}

    def handle_line(line):
        nonlocal result
        try:
            payload = json.loads(line)
        except Exception:
            return
        if payload.get("type") == "progress":
            if progress_callback:
                progress_callback(
                    payload.get("percent") or 0,
                    payload.get("message") or "正在识别视频文案",
                    payload,
                )
        elif "ok" in payload:
            result = payload

    returncode, output = run_streaming_command(command, timeout=1800, on_line=handle_line)
    if returncode != 0 or not result.get("ok"):
        detail = result.get("error") or "\n".join(output[-30:]) or "Whisper 转写失败"
        raise RuntimeError(detail[-2000:])
    return result["data"], model


def transcribe_idea_radar_media(media_path, progress_callback=None):
    return idea_radar_media.transcribe_media(
        media_path,
        progress_callback=progress_callback,
        log=backend_log,
    )


def download_idea_radar_media(video_url, work_dir, progress_callback=None):
    """Use the reusable video-jiexi adapter first, then keep the local fallback."""
    settings = load_runtime_settings()
    if video_jiexi_client.base_url(settings):
        try:
            if progress_callback:
                progress_callback(5, "正在调用视频解析服务获取公开媒体")
            inspection = video_jiexi_client.inspect(video_url, settings=settings)
            inspection_id = inspection.get("inspectionId") if isinstance(inspection, dict) else None
            if not inspection_id:
                raise RuntimeError("视频解析服务未返回 inspectionId")
            formats = inspection.get("formats") if isinstance(inspection, dict) else []
            format_id = str((formats or [{}])[0].get("id") or "")
            task = video_jiexi_client.start_download(inspection_id, format_id or None, "video", settings)
            task_id = task.get("id") if isinstance(task, dict) else None
            if not task_id:
                raise RuntimeError("视频解析服务未返回下载任务 ID")
            while True:
                current = video_jiexi_client.get_task(task_id, settings)
                state = str(current.get("state") or "")
                percent = float(current.get("progress") or 0)
                if progress_callback:
                    progress_callback(min(98, max(8, percent)), "视频解析服务正在下载")
                if state == "completed":
                    filename, content = video_jiexi_client.download_file(task_id, settings)
                    suffix = Path(filename or "").suffix or ".mp4"
                    target = Path(work_dir) / f"source{suffix}"
                    target.write_bytes(content)
                    if progress_callback:
                        progress_callback(100, "视频解析服务下载完成")
                    return target
                if state in {"error", "cancelled"}:
                    raise RuntimeError(current.get("error") or f"下载任务{state}")
                time.sleep(1.5)
        except Exception as exc:
            backend_log(f"video-jiexi idea radar download failed, fallback to direct: {exc}")
            if progress_callback:
                progress_callback(5, "视频解析服务暂不可用，切换备用下载方式")
    return download_douyin_video(
        video_url, work_dir,
        base_dir=BASE_DIR,
        latest_cookie_file=latest_douyin_cookie_file,
        progress_callback=progress_callback,
        log=backend_log,
    )


def clean_transcript_text(text):
    return idea_radar_media.clean_transcript_text(text)


def parse_metric_number(value):
    text = re.sub(r"\s+", "", str(value or "")).replace(",", "")
    if not text:
        return 0
    match = re.search(r"([\d.]+)", text)
    if not match:
        return 0
    number = float(match.group(1))
    if "万" in text or "w" in text.lower():
        number *= 10000
    return int(number)


def list_idea_radar_videos(limit=80, days=0):
    sync_result = sync_hot_monitor_works()
    active_monitor_urls = None if sync_result.get("error") else set(sync_result["active_urls"])
    rules = get_benchmark_monitoring_defaults()
    return benchmark_repository.list_idea_radar_videos(
        get_db_path(), parse_metric_number, limit, rules["hot_multiple"], days,
        active_monitor_urls,
    )


def sync_hot_monitor_works():
    """Mirror the current hot queue into the reusable analysis pipeline."""
    try:
        works = opencli_monitor_service.list_analysis_queue(load_runtime_settings())
    except Exception as exc:
        backend_log(f"hot queue mirror skipped: {exc}")
        return {"synced": 0, "skipped": 0, "error": str(exc)}
    synced = 0
    skipped = 0
    active_urls = []
    for work in works:
        video_url = str(work.get("url") or work.get("video_url") or "").strip()
        if video_url:
            active_urls.append(video_url)
        if benchmark_repository.upsert_monitor_queue_video(get_db_path(), work):
            synced += 1
        else:
            skipped += 1
    return {"synced": synced, "skipped": skipped, "active_urls": active_urls}


def classify_idea_theme(text):
    checks = [
        (["变现", "赚钱", "收入", "交付", "获客", "会员费", "分账"], "AI 变现闭环"),
        (["Skill", "智能体", "Agent", "工作流", "自审", "预测", "蒸馏"], "AI 生产系统"),
        (["开源", "Github", "GitHub", "项目", "CLI", "工具"], "开源工具机会"),
        (["第二大脑", "知识库", "聊天记录", "备份", "数字资产"], "数字资产沉淀"),
        (["入门", "小白", "教程", "学习", "电脑", "配置"], "普通人入门路径"),
        (["职业", "工程师", "GEO", "优化", "部署"], "AI 新职业地图"),
    ]
    for keywords, label in checks:
        if any(keyword.lower() in text.lower() for keyword in keywords):
            return label
    return "AI 机会识别"


def pick_evidence_type(text, like_count):
    types = []
    if parse_metric_number(like_count) >= 10000:
        types.append("高赞验证")
    if any(word in text for word in ["开源", "公开", "收入公开"]):
        types.append("公开证据")
    if any(word in text for word in ["全链路", "拆解", "教程", "指南"]):
        types.append("方法拆解")
    if any(word in text for word in ["我", "自用", "实测", "体验"]):
        types.append("亲身实验")
    return types or ["标题信号"]


def build_migration_angles(theme, text):
    if theme == "AI 变现闭环":
        return [
            "把 AI 热点翻译成普通人能执行的赚钱链路",
            "从工具介绍升级成开发、获客、交付的闭环拆解",
            "用自己的发布和复盘数据证明内容系统能产生反馈",
        ]
    if theme == "AI 生产系统":
        return [
            "把单个工具改写成一套可复用 Skill 资产",
            "用对标研究、观点提炼、发布复盘组成个人内容流水线",
            "展示一个真实任务如何被 AI 工作流接管",
        ]
    if theme == "开源工具机会":
        return [
            "把开源项目从技术新闻改写成普通人的机会窗口",
            "用项目增长速度、使用场景和替代对象讲清趋势",
            "输出一条从发现项目到内容选题的研究路径",
        ]
    if theme == "数字资产沉淀":
        return [
            "把聊天记录、选题、复盘沉淀成未来可调用资产",
            "证明数据不是看板，而是下一条内容的原料",
            "从资料收藏升级成第二大脑的运行机制",
        ]
    if theme == "普通人入门路径":
        return [
            "把入门问题拆成最低成本的第一步",
            "反对先买设备、先报课，强调先跑通一个真实任务",
            "用对标账号样本生成一条新手路线图",
        ]
    return [
        "把抽象趋势改写成普通人的具体行动",
        "从对标账号中提炼可迁移观点，而不是复制标题",
        "把热点变成自己的长期栏目和方法论资产",
    ]


def generate_idea_radar(video, target_direction=None, transcript=None):
    """Platform-agnostic local fallback when the optional AI layer is unavailable."""
    account = video.get("account_name") or "对标账号"
    title = str(video.get("title") or "这条对标作品").strip()
    topic = re.sub(r"\s*#.*$", "", title).strip()[:42] or "这条对标作品"
    transcript = str(transcript or "").strip()
    like_count = video.get("like_count") or "0"
    metrics = [
        f"点赞 {like_count}",
        f"评论 {video.get('comment_count') or '未提供'}",
        f"收藏 {video.get('collect_count') or '未提供'}",
        f"分享 {video.get('share_count') or '未提供'}",
    ]
    is_food = any(word in f"{title} {transcript}" for word in ["汉堡", "冰淇淋", "可乐", "美食", "口味", "好吃", "餐"])
    theme = "产品实测与性价比" if is_food else "具体场景下的经验分享"
    hook = split_title_parts(title)[0][:48] or topic
    viewpoint = (
        "观众愿意停下来，不一定是为了产品本身，而是想尽快知道它值不值、怎么选、有没有坑。"
        if is_food else "观众愿意停下来，不只是因为话题，而是想获得一个更省时间、更能执行的判断。"
    )
    titles = [
        f"{topic}到底值不值？我只讲三个实测判断",
        f"别只看宣传：{topic}真正该怎么选",
        f"把{topic}拆开看，最容易被忽略的是这一步",
    ]
    variants = [
        {
            "level": "轻度改编", "title": titles[0],
            "what_to_keep": "保留原作的具体对象、开头反差和快速给判断的节奏。",
            "what_to_change": "换成自己的实测对象、镜头和结论依据，不复述原作者体验。",
            "script_outline": f"前三秒抛出“{hook}”；展示一个关键细节；给出值不值的结论；补一个适合/不适合的人群。",
        },
        {
            "level": "中度改编", "title": titles[1],
            "what_to_keep": "保留“替观众做选择”的实用价值。",
            "what_to_change": "改用对比叙事：原本以为怎样，实测后为什么改变判断。",
            "script_outline": "先给预期，再给实测反差；用两到三个具体标准解释；最后给明确选择建议。",
        },
        {
            "level": "深度改编", "title": titles[2],
            "what_to_keep": "保留真实场景和可讨论的观点。",
            "what_to_change": "从单品评测扩展为一套选择方法，增加反例与限制条件。",
            "script_outline": "用一个常见误区开场；拆解判断标准；给一个反例；让观众在评论区补充自己的标准。",
        },
    ]
    script = (
        f"{hook}。\n\n"
        f"我这次不急着下结论，先把大家最在意的三个点讲清楚。第一，看它到底有没有解决实际场景里的问题；第二，看同价位有没有更稳的选择；第三，看它适不适合你自己的需求。\n\n"
        f"原作品之所以容易被讨论，不只是因为{topic}，而是它把一个模糊的“值不值”变成了能立刻判断的细节。\n\n"
        "所以如果我重新拍，我会保留这个判断框架，但换成自己的实测、自己的对比和自己的结论。别急着跟风，先看你真正需要的是什么。你会怎么选，评论区说说。"
    )
    return {
        "radar_type": "metadata", "target_direction": "",
        "source": {"id": video.get("id"), "account_name": account, "title": title, "video_url": video.get("video_url"), "like_count": like_count, "like_score": parse_metric_number(like_count), "cover_url": video.get("cover_url")},
        "viral_theme": theme,
        "audience_anxieties": ["不想花冤枉时间或钱", "想快速得到明确判断", "担心只看到宣传、看不到真实体验"],
        "contrarian_viewpoint": viewpoint,
        "evidence_types": ["公开互动信号", "视频转写" if transcript else "标题与公开字段", "编辑推断", "待验证假设"],
        "migration_angles": ["保留具体场景，换成自己的实测", "保留判断框架，补充不同对比对象", "把单点体验升级为可复用选择方法"],
        "recommended_titles": titles,
        "opening_script": f"{hook}。别急着跟风，我先把最关键的判断点讲清楚。",
        "personalized_script": script, "complete_script": script, "adaptation_variants": variants,
        "formula": "可参考性 = 具体场景 × 明确判断 × 真实细节 × 可执行选择",
        "content_breakdown": {
            "summary": f"作品围绕“{topic}”建立具体场景，再用实测细节帮助观众做选择。",
            "target_audience": "对同类产品或场景有直接需求、想快速判断值不值的人。",
            "hook": hook, "structure": ["具体对象或反差开头", "展示关键细节", "给出判断", "补充适用条件或互动引导"],
            "core_viewpoint": viewpoint, "evidence": metrics + (["本地转写已完成"] if transcript else []),
            "emotional_turn": "从好奇或犹豫，转向获得一个可执行的选择标准。",
            "spread_promise": "让观众快速获得一个能复述、能拿去判断的结论。",
            "reusable_mechanisms": ["具体对象开头", "先结论后依据", "明确适用与不适用人群"],
            "non_reusable_parts": ["原作者的个人体验和镜头素材", "未提供的评论样本、播放量和完播率"],
            "opportunity_chain": ["原作提供具体场景", "提炼判断标准", "换成自己的证据重新表达"],
            "gaps": ["公开页面通常没有播放量和完播率", "当前没有评论样本或关键帧证据"],
            "confidence": "中：热度、标题和转写可验证；传播因果仍需通过二创发布结果验证。"
        }
    }


def get_transcript_radar_schema():
    return transcript_radar_schema()


def build_transcript_radar_prompt(video, transcript, target_direction):
    return build_idea_radar_prompt(video, transcript, target_direction)


def run_codex_cli_structured(prompt, schema, timeout=None):
    timeout = timeout or int(get_runtime_setting(
        "codexAnalysisTimeout", "codex_timeout", env="CODEX_ANALYSIS_TIMEOUT_SECONDS", default="180"
    ))
    configured_codex_cmd = get_runtime_setting(
        "codexCliPath", "codex_cli_path", env="CODEX_CLI_PATH", default="codex"
    )
    codex_model = get_runtime_setting(
        "codexAnalysisModel", "codex_model", env="CODEX_ANALYSIS_MODEL", default="gpt-5.4-mini"
    )
    return agent_run_codex_cli_structured(
        prompt,
        schema,
        base_dir=BASE_DIR,
        codex_cmd=configured_codex_cmd,
        codex_model=codex_model,
        timeout=timeout,
        log=backend_log,
    )


def get_task_agent_model(task_name="viralAnalysis", settings=None):
    return agent_get_task_agent_model(task_name, settings or load_runtime_settings())


def call_hermes_structured(prompt, schema, model_config=None, timeout=None):
    settings = load_runtime_settings()
    return agent_call_hermes_structured(
        prompt,
        schema,
        settings=settings,
        model_config=model_config,
        timeout=timeout,
        log=backend_log,
    )


def run_agent_structured(prompt, schema, model_config=None, timeout=None, task_name="viralAnalysis"):
    settings = load_runtime_settings()
    if model_config is None:
        ensure_default_agent_model()
        settings = load_runtime_settings()
        model_config = get_task_agent_model(task_name, settings)
    provider = str(model_config.get("provider") or "").strip().lower()
    if provider in {"codex", "codex-cli", "openai-codex-cli"}:
        configured_codex_cmd = get_runtime_setting(
            "codexCliPath", "codex_cli_path", env="CODEX_CLI_PATH",
            default="codex.cmd" if os.name == "nt" else "codex",
        )
        return agent_run_codex_cli_structured(
            prompt,
            schema,
            base_dir=BASE_DIR,
            codex_cmd=configured_codex_cmd,
            codex_model=model_config.get("model") or "gpt-5.4-mini",
            timeout=timeout or int(get_runtime_setting(
                "codexAnalysisTimeout", "codex_timeout",
                env="CODEX_ANALYSIS_TIMEOUT_SECONDS", default="300",
            )),
            log=backend_log,
        )
    if provider in {"universal-ai", "openai-compatible", "openai_compatible"}:
        return agent_call_universal_ai_structured(
            prompt, schema, settings=settings, model_config=model_config,
            timeout=timeout, log=backend_log,
        )
    return call_hermes_structured(
        prompt, schema, model_config=model_config, timeout=timeout
    )


def run_codex_structured(prompt, schema, timeout=None):
    return run_agent_structured(prompt, schema, timeout=timeout)


def generate_identity_script(identity_profile, radar_result, benchmark_analysis=None):
    return run_agent_structured(
        build_identity_script_prompt(identity_profile, radar_result, benchmark_analysis),
        identity_script_schema(),
        timeout=get_hermes_settings(load_runtime_settings())["timeout"],
        task_name="scriptGeneration",
    )


def load_idea_radar_video(video_id):
    return benchmark_repository.get_video(get_db_path(), video_id, include_account=True)


def run_idea_radar_pipeline(video_id, target_direction, force_transcription=False, transcribe_only=False):
    def run_structured_with_fallback(prompt, schema):
        try:
            return run_codex_structured(prompt, schema)
        except Exception as exc:
            # AI 是增强层，不应阻断已完成的视频下载、转写和基础拆解。
            video = load_idea_radar_video(video_id) or {}
            transcript_state = get_idea_radar_transcript(video_id) or {}
            fallback_video = dict(video)
            fallback = generate_idea_radar(
                fallback_video,
                target_direction,
                transcript=transcript_state.get("cleaned_transcript") or "",
            )
            fallback["ai_status"] = "unavailable"
            fallback["ai_error"] = str(exc)[:300]
            return fallback

    return run_idea_radar_pipeline_core(
        video_id,
        target_direction,
        force_transcription=force_transcription,
        transcribe_only=transcribe_only,
        load_video=load_idea_radar_video,
        get_transcript=get_idea_radar_transcript,
        update_progress=update_idea_radar_progress,
        download_video=download_idea_radar_media,
        transcribe_media=transcribe_idea_radar_media,
        clean_transcript=clean_transcript_text,
        build_prompt=build_transcript_radar_prompt,
        schema_factory=get_transcript_radar_schema,
        run_structured=run_structured_with_fallback,
        get_agent_model=get_task_agent_model,
        parse_metric_number=parse_metric_number,
        registry=idea_radar_job_registry,
    )


def start_idea_radar_pipeline(video_id, target_direction, force=False, force_transcription=False, transcribe_only=False):
    current = get_idea_radar_transcript(video_id)
    started = start_pipeline_task(
        video_id,
        target_direction,
        current=current,
        registry=idea_radar_job_registry,
        update_transcript=update_idea_radar_transcript,
        pipeline_fn=run_idea_radar_pipeline,
        force=force,
        force_transcription=force_transcription,
        transcribe_only=transcribe_only,
    )
    return started or get_idea_radar_transcript(video_id)


ensure_douyin_benchmark_tables()
ensure_douyin_own_tables()
ensure_xiaohongshu_own_tables()

# 获取当前目录。开发环境由 Vite 提供前端，打包环境优先使用前端构建产物。
current_dir = os.path.dirname(os.path.abspath(__file__))
frontend_dist_dir = os.path.join(current_dir, 'sau_frontend', 'dist')

# 处理打包后的前端资源请求，同时保留旧版 assets 目录兼容性。
@app.route('/assets/<path:filename>')
def custom_static(filename):
    dist_assets_dir = os.path.join(frontend_dist_dir, 'assets')
    if os.path.isfile(os.path.join(dist_assets_dir, filename)):
        return send_from_directory(dist_assets_dir, filename)
    return send_from_directory(os.path.join(current_dir, 'assets'), filename)

# 处理 favicon.ico 静态资源（未来打包用）
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(current_dir, 'assets'), 'favicon.ico')

# 打包模式直接提供 Vite 构建产物；未构建前返回可操作的健康信息，避免 500。
@app.route('/')
def hello_world():  # put application's code here
    frontend_index = os.path.join(frontend_dist_dir, 'index.html')
    if os.path.isfile(frontend_index):
        return send_from_directory(frontend_dist_dir, 'index.html')
    return jsonify({
        "code": 200,
        "msg": "Content workbench backend is online. Run npm.cmd run build in sau_frontend to serve the UI.",
        "data": {"frontend": "not_built"},
    })

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({
            "code": 200,
            "data": None,
            "msg": "No file part in the request"
        }), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({
            "code": 200,
            "data": None,
            "msg": "No selected file"
        }), 400
    try:
        generated = material_service.save_temp_upload(file, BASE_DIR)
        return jsonify({"code":200,"msg": "File uploaded successfully", "data": generated}), 200
    except Exception as e:
        return jsonify({"code":200,"msg": str(e),"data":None}), 500

@app.route('/getFile', methods=['GET'])
def get_file():
    filename = request.args.get('filename')

    try:
        filename = material_service.sanitize_download_filename(filename)
    except ValueError as exc:
        return {"error": str(exc)}, 400
    return send_from_directory(str(material_service.material_dir(BASE_DIR)), filename)


@app.route('/uploadSave', methods=['POST'])
def upload_save():
    if 'file' not in request.files:
        return jsonify({
            "code": 400,
            "data": None,
            "msg": "No file part in the request"
        }), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({
            "code": 400,
            "data": None,
            "msg": "No selected file"
        }), 400

    # 获取表单中的自定义文件名（可选）
    custom_filename = request.form.get('filename', None)
    if custom_filename:
        filename = custom_filename + "." + file.filename.split('.')[-1]
    else:
        filename = file.filename

    try:
        saved = material_service.save_material_upload(
            get_db_path(),
            file,
            BASE_DIR,
            custom_filename=custom_filename,
        )
        print("[OK] 上传文件已记录")

        return jsonify({
            "code": 200,
            "msg": "File uploaded and saved successfully",
            "data": saved
        }), 200

    except Exception as e:
        return jsonify({
            "code": 500,
            "msg": str("upload failed!"),
            "data": None
        }), 500

@app.route('/getFiles', methods=['GET'])
def get_all_files():
    try:
        data = material_repository.list_file_records(get_db_path())
        return jsonify({
            "code": 200,
            "msg": "success",
            "data": data
        }), 200
    except Exception as e:
        return jsonify({
            "code": 500,
            "msg": str("get file failed!"),
            "data": None
        }), 500


@app.route('/account/updateFollower', methods=['POST'])
def update_account_follower():
    """更新账号粉丝数"""
    try:
        data = request.get_json()
        account_id = data.get('id')
        follower_count = data.get('follower_count')
        if not account_id or follower_count is None:
            return jsonify({"code": 400, "msg": "缺少账号ID或粉丝数"}), 400

        account_repository.update_follower_count(get_db_path(), account_id, follower_count)

        return jsonify({"code": 200, "msg": "更新成功"}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/account/followers', methods=['GET'])
def get_account_followers():
    """获取所有账号的粉丝数"""
    try:
        accounts = account_repository.list_followers(get_db_path())
        return jsonify({"code": 200, "msg": "success", "data": accounts}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/accounts', methods=['POST'])
def add_douyin_benchmark_account():
    data = request.get_json() or {}
    homepage_url = data.get("homepageUrl") or data.get("homepage_url")
    try:
        cookie_file = latest_douyin_cookie_file()
        result = sync_one_douyin_benchmark(homepage_url, cookie_file, max_videos=20)
        return jsonify({"code": 200, "msg": "success", "data": {"id": result["id"], "sync": result["sync"]}}), 200
    except Exception as e:
        try:
            if homepage_url:
                save_douyin_benchmark_error(upsert_douyin_benchmark(homepage_url), str(e))
        except Exception:
            pass
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/monitor/accounts', methods=['POST'])
def bind_opencli_monitor_account():
    """Register a platform-agnostic benchmark account in OpenCLI Admin."""
    payload = request.get_json(silent=True) or {}
    platform = str(payload.get("platform") or "douyin").strip().lower()
    account_reference = (
        payload.get("accountRef")
        or payload.get("account_ref")
        or payload.get("homepageUrl")
        or payload.get("homepage_url")
        or payload.get("external_account_id")
        or payload.get("sec_uid")
    )
    try:
        settings = load_runtime_settings()
        result = opencli_monitor_service.bind_account(
            platform,
            account_reference,
            settings,
            monitoring_rules=get_benchmark_monitoring_defaults(settings),
        )
        return jsonify({"code": 200, "msg": "success", "data": result}), 201
    except ValueError as exc:
        return jsonify({"code": 400, "msg": str(exc), "data": None}), 400
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/rules', methods=['GET'])
def get_benchmark_monitor_rules():
    return jsonify({"code": 200, "msg": "success", "data": get_benchmark_monitoring_defaults()}), 200


@app.route('/benchmark/monitor/rules', methods=['PUT'])
def update_benchmark_monitor_rules():
    payload = request.get_json(silent=True) or {}
    try:
        rules = normalize_benchmark_monitoring_rules(payload, inherit_global=True)
        rules["inherit_global"] = True
    except (TypeError, ValueError) as exc:
        return jsonify({"code": 400, "msg": str(exc), "data": None}), 400
    settings = load_runtime_settings()
    settings["benchmarkMonitoringDefaults"] = rules
    save_runtime_settings(settings)
    applied = 0
    failures = []
    try:
        accounts = opencli_monitor_service.list_accounts(None, settings)
        for account in accounts:
            if account.get("collection_enabled") is False and account.get("collection_status") != "paused":
                continue
            current = account.get("monitoring_rules") or {}
            if current.get("inherit_global", True) is not True:
                continue
            try:
                opencli_monitor_service.update_account(
                    account["id"], {"monitoring_rules": rules}, settings
                )
                applied += 1
            except opencli_monitor_service.OpenCLIAdminError as exc:
                failures.append({"account_id": account.get("id"), "error": str(exc)})
    except opencli_monitor_service.OpenCLIAdminError as exc:
        failures.append({"account_id": None, "error": str(exc)})
    return jsonify({
        "code": 200,
        "msg": "success" if not failures else "默认规则已保存，部分账号同步失败",
        "data": {"rules": rules, "accounts_updated": applied, "failures": failures},
    }), 200


@app.route('/benchmark/platforms', methods=['GET'])
def list_benchmark_platforms():
    """List platform adapters configured in the optional collection service."""
    try:
        platforms = opencli_monitor_service.list_platforms(load_runtime_settings())
        return jsonify({"code": 200, "msg": "success", "data": platforms}), 200
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/accounts', methods=['GET'])
def list_opencli_monitor_accounts():
    try:
        platform = request.args.get("platform") or None
        accounts = opencli_monitor_service.list_accounts(
            platform, load_runtime_settings()
        )
        for account in accounts:
            account["monitoring_rules"] = normalize_benchmark_monitoring_rules(
                account.get("monitoring_rules"), inherit_global=True
            )
        return jsonify({"code": 200, "msg": "success", "data": accounts}), 200
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/accounts/<account_id>/check', methods=['POST'])
def check_opencli_monitor_account(account_id):
    try:
        result = opencli_monitor_service.check_account(account_id, load_runtime_settings())
        return jsonify({"code": 202, "msg": "success", "data": result}), 202
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/accounts/<account_id>', methods=['DELETE'])
def remove_opencli_monitor_account(account_id):
    try:
        result = opencli_monitor_service.remove_account(
            account_id, settings=load_runtime_settings()
        )
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/accounts/<account_id>', methods=['PATCH'])
def update_opencli_monitor_account(account_id):
    payload = request.get_json(silent=True) or {}
    changes = {}
    display_name = str(payload.get("displayName") or payload.get("display_name") or "").strip()
    if display_name:
        changes["display_name"] = display_name
    if "monitoringRules" in payload or "monitoring_rules" in payload:
        try:
            changes["monitoring_rules"] = normalize_benchmark_monitoring_rules(
                payload.get("monitoringRules") or payload.get("monitoring_rules"),
                inherit_global=False,
            )
        except (TypeError, ValueError) as exc:
            return jsonify({"code": 400, "msg": str(exc), "data": None}), 400
    if "enabled" in payload:
        changes["enabled"] = bool(payload.get("enabled"))
    if not changes:
        return jsonify({"code": 400, "msg": "没有可保存的账号设置", "data": None}), 400
    try:
        result = opencli_monitor_service.update_account(
            account_id, changes, load_runtime_settings()
        )
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/benchmark/monitor/works', methods=['GET'])
def list_opencli_monitor_works():
    try:
        platform = request.args.get("platform") or None
        works = opencli_monitor_service.list_analysis_queue(
            load_runtime_settings(), platform=platform
        )
        return jsonify({"code": 200, "msg": "success", "data": works}), 200
    except opencli_monitor_service.OpenCLIAdminError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/integrations/video-jiexi/status', methods=['GET'])
def video_jiexi_status():
    """Return health information for the bundled parser service."""
    settings = load_runtime_settings()
    configured_url = video_jiexi_client.base_url(settings)
    if not configured_url:
        return jsonify({"code": 200, "msg": "success", "data": {"configured": False, "available": False, "base_url": ""}}), 200
    try:
        health = video_jiexi_client.health(settings)
        embedded = configured_url.startswith("http://127.0.0.1:") and configured_url.rsplit(":", 1)[-1] != "4200"
        return jsonify({"code": 200, "msg": "success", "data": {"configured": True, "embedded": embedded, "base_url": "" if embedded else configured_url, "download_dir": str(video_jiexi_client.download_root(settings) or health.get("downloadDir") or ""), "health": health}}), 200
    except video_jiexi_client.VideoJiexiError as exc:
        return jsonify({"code": 200, "msg": "success", "data": {"configured": True, "embedded": configured_url.startswith("http://127.0.0.1:"), "base_url": "", "available": False, "error": str(exc)}}), 200


@app.route('/integrations/video-jiexi/open-folder', methods=['POST'])
def video_jiexi_open_folder():
    """Open the local video-jiexi download directory in File Explorer."""
    settings = load_runtime_settings()
    folder = video_jiexi_client.download_root(settings)
    if folder is None:
        return jsonify({"code": 400, "msg": "未配置视频下载目录", "data": None}), 400
    folder = folder.expanduser().resolve()
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
    try:
        # Launch Explorer as a separate foreground shell.  This is more
        # reliable than os.startfile when the workbench is running behind a
        # browser window and the user explicitly asked to open the folder.
        if os.name == 'nt':
            subprocess.Popen(['explorer.exe', str(folder)], close_fds=True)
        else:
            os.startfile(str(folder))
    except (AttributeError, OSError) as exc:
        return jsonify({"code": 500, "msg": f"无法打开下载目录：{exc}", "data": {"path": str(folder)}}), 500
    return jsonify({"code": 200, "msg": "success", "data": {"path": str(folder)}}), 200


@app.route('/integrations/video-jiexi/inspect', methods=['POST'])
def video_jiexi_inspect():
    payload = request.get_json(silent=True) or {}
    url = str(payload.get('url') or '').strip()
    if not url:
        return jsonify({"code": 400, "msg": "请提供视频链接", "data": None}), 400
    try:
        result = video_jiexi_client.inspect(url, payload.get('cookieBrowser') or payload.get('cookie_browser') or '', load_runtime_settings())
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except video_jiexi_client.VideoJiexiError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/integrations/video-jiexi/download', methods=['POST'])
def video_jiexi_download():
    payload = request.get_json(silent=True) or {}
    inspection_id = str(payload.get('inspectionId') or payload.get('inspection_id') or '').strip()
    if not inspection_id:
        return jsonify({"code": 400, "msg": "缺少解析结果 ID", "data": None}), 400
    try:
        result = video_jiexi_client.start_download(
            inspection_id,
            payload.get('formatId') or payload.get('format_id'),
            payload.get('kind') or 'video',
            load_runtime_settings(),
        )
        return jsonify({"code": 202, "msg": "success", "data": result}), 202
    except video_jiexi_client.VideoJiexiError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/integrations/video-jiexi/tasks/<task_id>', methods=['GET'])
def video_jiexi_task(task_id):
    try:
        result = video_jiexi_client.get_task(task_id, load_runtime_settings())
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except video_jiexi_client.VideoJiexiError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/integrations/video-jiexi/import', methods=['POST'])
def video_jiexi_import():
    payload = request.get_json(silent=True) or {}
    task_id = str(payload.get('taskId') or payload.get('task_id') or '').strip()
    if not task_id:
        return jsonify({"code": 400, "msg": "缺少下载任务 ID", "data": None}), 400
    try:
        settings = load_runtime_settings()
        task = video_jiexi_client.get_task(task_id, settings)
        if task.get('state') != 'completed':
            return jsonify({"code": 409, "msg": "下载任务尚未完成", "data": task}), 409
        try:
            original_name, raw_content = video_jiexi_client.download_file(task_id, settings)
        except video_jiexi_client.VideoJiexiError:
            source = video_jiexi_client.task_file_path(task, settings)
            original_name = source.name
            raw_content = source.read_bytes()
        target_name = f"{uuid.uuid1()}_{original_name}"
        target = material_service.material_dir(BASE_DIR) / target_name
        target.write_bytes(raw_content)
        filesize = round(float(target.stat().st_size) / (1024 * 1024), 2)
        material_repository.add_file_record(get_db_path(), original_name, filesize, target_name)
        return jsonify({"code": 200, "msg": "文件已保存到本地素材目录", "data": {"filename": original_name, "filepath": target_name, "filesize": filesize}}), 200
    except video_jiexi_client.VideoJiexiError as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502
    except OSError as exc:
        return jsonify({"code": 500, "msg": f"导入素材失败：{exc}", "data": None}), 500


@app.route('/benchmark/douyin/auto-discover', methods=['POST'])
def auto_discover_douyin_benchmark_accounts():
    payload = request.get_json() or {}
    keywords = parse_keywords_payload(payload.get("keywords") or payload.get("keyword"))
    try:
        if not keywords:
            return jsonify({"code": 400, "msg": "请填写至少一个对标关键词", "data": None}), 400
        cookie_file = latest_douyin_cookie_file()
        limit = max(1, min(int(payload.get("limit") or 5), 20))
        max_videos = max(1, min(int(payload.get("maxVideos") or payload.get("max_videos") or 10), 30))
        candidates = asyncio.run(discover_douyin_benchmark_accounts(
            keywords,
            cookie_file=cookie_file,
            limit=limit
        ))
        synced = []
        failed = []
        processed_urls = set()
        for candidate in candidates:
            homepage_url = normalize_douyin_user_url(candidate.get("homepage_url"))
            if not homepage_url or homepage_url in processed_urls:
                continue
            processed_urls.add(homepage_url)
            candidate["homepage_url"] = homepage_url
            try:
                result = sync_one_douyin_benchmark(homepage_url, cookie_file, max_videos=max_videos)
                synced.append({
                    "candidate": candidate,
                    "id": result["id"],
                    "sync": result["sync"],
                    "nickname": result["account"].get("nickname") or candidate.get("nickname"),
                    "homepage_url": homepage_url,
                })
            except Exception as exc:
                try:
                    if homepage_url:
                        save_douyin_benchmark_error(upsert_douyin_benchmark(homepage_url), str(exc))
                except Exception:
                    pass
                failed.append({
                    "candidate": candidate,
                    "homepage_url": homepage_url,
                    "error": str(exc),
                })
        return jsonify({
            "code": 200,
            "msg": "success",
            "data": {
                "keywords": keywords,
                "candidates": candidates,
                "synced": synced,
                "failed": failed,
                "summary": {
                    "found": len(candidates),
                    "synced": len(synced),
                    "failed": len(failed),
                }
            }
        }), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/douyin/import/preview', methods=['POST'])
def preview_own_douyin_import():
    try:
        if 'file' not in request.files:
            return jsonify({"code": 400, "msg": "请上传 CSV 或 XLSX 文件", "data": None}), 400
        parsed = parse_own_douyin_import(request.files['file'])
        return jsonify({
            "code": 200,
            "msg": "success",
            "data": {
                "headers": parsed["headers"],
                "field_map": parsed["field_map"],
                "raw_count": parsed["raw_count"],
                "valid_count": parsed["valid_count"],
                "preview_rows": parsed["rows"][:10],
            }
        }), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/douyin/import', methods=['POST'])
def import_own_douyin_videos():
    try:
        if 'file' not in request.files:
            return jsonify({"code": 400, "msg": "请上传 CSV 或 XLSX 文件", "data": None}), 400
        account_name = request.form.get("accountName") or request.form.get("account_name") or "我的账号"
        parsed = parse_own_douyin_import(request.files['file'])
        result = save_own_douyin_import(parsed["rows"], account_name=account_name)
        result.update({
            "raw_count": parsed["raw_count"],
            "valid_count": parsed["valid_count"],
            "field_map": parsed["field_map"],
        })
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/douyin/videos', methods=['GET'])
def get_own_douyin_videos():
    try:
        limit = request.args.get("limit", 100)
        videos = list_own_douyin_videos(limit)
        return jsonify({"code": 200, "msg": "success", "data": videos}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/douyin/overview', methods=['GET'])
def get_own_douyin_account_overview():
    try:
        return jsonify({"code": 200, "msg": "success", "data": get_own_douyin_overview()}), 200
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/own/xiaohongshu/import/preview', methods=['POST'])
def preview_own_xiaohongshu_import():
    """Preview normalized Xiaohongshu creator exports using the shared schema."""
    try:
        if 'file' not in request.files:
            return jsonify({"code": 400, "msg": "请上传 CSV 或 XLSX 文件", "data": None}), 400
        parsed = parse_own_douyin_import(request.files['file'])
        return jsonify({
            "code": 200, "msg": "success",
            "data": {
                "headers": parsed["headers"], "field_map": parsed["field_map"],
                "raw_count": parsed["raw_count"], "valid_count": parsed["valid_count"],
                "preview_rows": parsed["rows"][:10],
            },
        }), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/xiaohongshu/import', methods=['POST'])
def import_own_xiaohongshu_videos():
    try:
        if 'file' not in request.files:
            return jsonify({"code": 400, "msg": "请上传 CSV 或 XLSX 文件", "data": None}), 400
        account_name = (
            request.form.get("accountName")
            or request.form.get("account_name")
            or "我的小红书账号"
        )
        parsed = parse_own_douyin_import(request.files['file'])
        result = save_own_xiaohongshu_import(parsed["rows"], account_name=account_name)
        result.update({
            "raw_count": parsed["raw_count"], "valid_count": parsed["valid_count"],
            "field_map": parsed["field_map"],
        })
        return jsonify({"code": 200, "msg": "success", "data": result}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/xiaohongshu/videos', methods=['GET'])
def get_own_xiaohongshu_videos():
    try:
        limit = request.args.get("limit", 100)
        account_id = request.args.get("account_id", type=int)
        videos = list_own_xiaohongshu_videos(limit, account_id=account_id)
        return jsonify({"code": 200, "msg": "success", "data": videos}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/own/xiaohongshu/overview', methods=['GET'])
def get_own_xiaohongshu_account_overview():
    try:
        account_id = request.args.get("account_id", type=int)
        return jsonify({"code": 200, "msg": "success", "data": get_own_xiaohongshu_overview(account_id=account_id)}), 200
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/own/<platform>/sync', methods=['POST'])
def sync_own_content(platform):
    """Synchronize creator data through the locally installed connector."""
    payload = request.get_json(silent=True) or {}
    account_name = str(payload.get("accountName") or "").strip()
    try:
        limit = max(1, min(int(payload.get("limit") or 20), 100))
        if platform == "douyin":
            result = own_content_connectors.sync_douyin(
                get_db_path(), account_name or "抖音创作者中心", recent_limit=limit
            )
        elif platform == "xiaohongshu":
            result = own_content_connectors.sync_xiaohongshu(
                get_db_path(), account_name or "我的小红书账号", limit=limit
            )
        else:
            return jsonify({"code": 400, "msg": "暂不支持该平台同步", "data": None}), 400
        return jsonify({"code": 200, "msg": "同步完成", "data": result}), 200
    except subprocess.TimeoutExpired:
        return jsonify({"code": 504, "msg": "平台连接器超时，请检查登录状态后重试", "data": None}), 504
    except Exception as exc:
        return jsonify({"code": 502, "msg": str(exc), "data": None}), 502


@app.route('/own/review/sources', methods=['GET'])
def get_own_review_sources():
    """Expose truthful connector state; no source is reported as connected by default."""
    connector_status = own_content_connectors.connector_availability()
    douyin_available = connector_status["douyin"].get("available", False)
    xhs_available = connector_status["xiaohongshu"].get("available", False)
    return jsonify({"code": 200, "msg": "success", "data": {
        "douyin": {
            "label": "抖音作品复盘", "connector": "Kuhakucai/douyin-mcp",
            "status": "sync_available" if douyin_available else "manual_import", "supports": [
                "曝光", "播放", "封面点击率", "平均播放时长", "完播率", "5秒完播率", "2秒跳出率",
                "推荐页", "搜索", "关注页", "个人主页", "粉丝/非粉丝占比", "作品主页访问",
                "粉丝趋势", "性别/年龄/地域画像", "点赞", "评论", "收藏", "分享",
            ],
            "note": ("以自然视频流量为主：优先展示曝光、点击、留存、自然来源和观众构成；DOU+、同城、好友与站外流量不进入核心展示。"
                     if douyin_available else connector_status["douyin"].get("error")),
        },
        "xiaohongshu": {
            "label": "小红书作品复盘", "connector": "OpenCLI 小红书创作者中心",
            "status": "sync_available" if xhs_available else "manual_import", "supports": [
                "观看", "点赞", "收藏", "评论", "分享", "涨粉", "观看来源", "观众画像", "账号趋势",
            ],
            "note": ("已接入小红书创作者中心：同步笔记明细、观看来源、观众画像，以及账号近30天趋势；平台暂未生成的指标保持为空。"
                     if xhs_available else connector_status["xiaohongshu"].get("error")),
        },
    }}), 200


@app.route('/platform-connections', methods=['GET'])
def get_platform_connections():
    """Return local connector and login state without exposing credentials."""
    try:
        probe = str(request.args.get("probe", "1")).strip().lower() not in {
            "0", "false", "no", "off"
        }
        data = platform_connections.all_connection_statuses(get_db_path(), probe=probe)
        return jsonify({"code": 200, "msg": "success", "data": data}), 200
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/platform-connections/<platform>/login', methods=['POST'])
def start_platform_login(platform):
    """Open the connector's visible official-site login flow on this computer."""
    payload = request.get_json(silent=True) or {}
    try:
        limit = max(5, min(int(payload.get("limit") or 20), 50))
        job = platform_connections.start_login(
            platform,
            get_db_path(),
            acknowledged_risk=bool(payload.get("acknowledgedRisk")),
            auto_sync=bool(payload.get("autoSync", True)),
            sync_limit=limit,
        )
        return jsonify({"code": 202, "msg": "已打开平台登录流程", "data": job}), 202
    except ValueError as exc:
        return jsonify({"code": 400, "msg": str(exc), "data": None}), 400
    except RuntimeError as exc:
        return jsonify({"code": 409, "msg": str(exc), "data": None}), 409
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/benchmark/douyin/accounts', methods=['GET'])
def get_douyin_benchmark_accounts():
    try:
        accounts = benchmark_repository.list_accounts(get_db_path())
        return jsonify({"code": 200, "msg": "success", "data": accounts}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/accounts/<int:account_id>', methods=['DELETE'])
def delete_douyin_benchmark_account(account_id):
    try:
        result = delete_douyin_benchmark(account_id)
        if not result:
            return jsonify({"code": 404, "msg": "对标账号不存在", "data": None}), 404
        return jsonify({
            "code": 200,
            "msg": "删除成功",
            "data": result,
        }), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/accounts/<int:account_id>/sync', methods=['POST'])
def sync_douyin_benchmark_account(account_id):
    try:
        cookie_file = latest_douyin_cookie_file()
        homepage_url = benchmark_repository.get_account_homepage(get_db_path(), account_id)
        if not homepage_url:
            return jsonify({"code": 404, "msg": "对标账号不存在", "data": None}), 404
        existing_urls = get_douyin_benchmark_video_urls(account_id)
        sync_data = asyncio.run(scrape_douyin_benchmark(
            homepage_url,
            cookie_file=cookie_file,
            max_videos=20,
            existing_video_urls=existing_urls
        ))
        stats = save_douyin_benchmark_sync(account_id, sync_data)
        return jsonify({"code": 200, "msg": "success", "data": {"id": account_id, "sync": stats}}), 200
    except Exception as e:
        save_douyin_benchmark_error(account_id, str(e))
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/accounts/<int:account_id>/videos', methods=['GET'])
def get_douyin_benchmark_videos(account_id):
    try:
        videos = benchmark_repository.list_videos(get_db_path(), account_id)
        return jsonify({"code": 200, "msg": "success", "data": videos}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/videos/<int:video_id>/analysis', methods=['GET'])
def get_douyin_benchmark_video_analysis(video_id):
    try:
        video = benchmark_repository.get_video(get_db_path(), video_id)
        if not video:
            return jsonify({"code": 404, "msg": "作品不存在", "data": None}), 404
        row = benchmark_repository.get_analysis(get_db_path(), video_id)
        if is_douyin_analysis_stale(row, video):
            analysis = generate_douyin_video_analysis(video)
            save_douyin_video_analysis(video_id, analysis)
            row = benchmark_repository.get_analysis(get_db_path(), video_id)
        return jsonify({
            "code": 200,
            "msg": "success",
            "data": {
                "video": video,
                "analysis": parse_douyin_video_analysis(row)
            }
        }), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/benchmark/douyin/videos/<int:video_id>/analysis', methods=['POST'])
def create_douyin_benchmark_video_analysis(video_id):
    try:
        force = bool((request.get_json(silent=True) or {}).get("force"))
        video = benchmark_repository.get_video(get_db_path(), video_id)
        if not video:
            return jsonify({"code": 404, "msg": "作品不存在", "data": None}), 404
        existing = benchmark_repository.get_analysis(get_db_path(), video_id)
        if existing and not force and not is_douyin_analysis_stale(existing, video):
            return jsonify({
                "code": 200,
                "msg": "success",
                "data": parse_douyin_video_analysis(existing)
            }), 200

        analysis = generate_douyin_video_analysis(video)
        save_douyin_video_analysis(video_id, analysis)
        row = benchmark_repository.get_analysis(get_db_path(), video_id)
        return jsonify({"code": 200, "msg": "success", "data": parse_douyin_video_analysis(row)}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/idea-radar/douyin/videos', methods=['GET'])
def get_idea_radar_videos():
    try:
        limit = request.args.get("limit", 80)
        days = request.args.get("days", 0)
        videos = list_idea_radar_videos(limit, days)
        return jsonify({"code": 200, "msg": "success", "data": videos}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/idea-radar/douyin/videos/manual', methods=['POST'])
def add_manual_idea_radar_video():
    try:
        payload = request.get_json(silent=True) or {}
        def normalize_manual_url(value):
            value = str(value or '').strip()
            if not value:
                return ''
            if not value.startswith(('http://', 'https://')):
                value = f'https://{value}'
            return value.split('?')[0].split('#')[0]
        video_url = str(payload.get('videoUrl') or payload.get('video_url') or '').strip()
        video_id = benchmark_repository.add_manual_video(get_db_path(), video_url, normalize_manual_url)
        inspect_error = None
        try:
            inspection = video_jiexi_client.inspect(normalize_manual_url(video_url), settings=load_runtime_settings())
            benchmark_repository.update_manual_video_metadata(get_db_path(), video_id, inspection)
        except Exception as exc:
            inspect_error = str(exc)[:300]
        return jsonify({"code": 200, "msg": "success", "data": load_idea_radar_video(video_id), "meta_error": inspect_error}), 200
    except ValueError as exc:
        return jsonify({"code": 400, "msg": str(exc), "data": None}), 400
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/idea-radar/douyin/videos/<int:video_id>/manual-details', methods=['PATCH'])
def update_manual_idea_radar_video_details(video_id):
    try:
        payload = request.get_json(silent=True) or {}
        ok = benchmark_repository.update_manual_video_details(
            get_db_path(), video_id,
            title=payload.get('title'), uploader=payload.get('uploader'), notes=payload.get('notes'),
        )
        if not ok:
            return jsonify({"code": 404, "msg": "手动作品不存在", "data": None}), 404
        return jsonify({"code": 200, "msg": "success", "data": load_idea_radar_video(video_id)}), 200
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/idea-radar/douyin/videos/<int:video_id>/analyze', methods=['POST'])
def analyze_idea_radar_video(video_id):
    try:
        payload = request.get_json(silent=True) or {}
        # 爆款拆解不依赖账号定位；保留空值只是兼容旧数据库字段。
        target_direction = ""
        if not load_idea_radar_video(video_id):
            return jsonify({"code": 404, "msg": "作品不存在", "data": None}), 404
        task = start_idea_radar_pipeline(
            video_id,
            target_direction,
            force=bool(payload.get("force")),
            force_transcription=bool(payload.get("forceTranscription") or payload.get("force_transcription")),
            transcribe_only=bool(load_runtime_settings().get("factsOnlyMode")),
        )
        return jsonify({"code": 200, "msg": "success", "data": task}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/idea-radar/douyin/videos/<int:video_id>/status', methods=['GET'])
def get_idea_radar_video_status(video_id):
    try:
        if not load_idea_radar_video(video_id):
            return jsonify({"code": 404, "msg": "作品不存在", "data": None}), 404
        task = get_idea_radar_status_payload(video_id, get_idea_radar_transcript)
        return jsonify({"code": 200, "msg": "success", "data": task}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


@app.route('/script-generation/identity', methods=['POST'])
def create_identity_script():
    try:
        payload = request.get_json(silent=True) or {}
        identity_profile = payload.get("identityProfile") or payload.get("identity_profile") or {}
        radar_result = payload.get("radarResult") or payload.get("radar_result") or {}
        benchmark_analysis = payload.get("benchmarkAnalysis") or payload.get("benchmark_analysis") or {}
        if not isinstance(identity_profile, dict) or not identity_profile:
            return jsonify({"code": 400, "msg": "请提供创作者身份资料", "data": None}), 400
        if not isinstance(radar_result, dict) or not radar_result:
            return jsonify({"code": 400, "msg": "请提供观点雷达结果", "data": None}), 400
        script = generate_identity_script(identity_profile, radar_result, benchmark_analysis)
        return jsonify({"code": 200, "msg": "success", "data": script}), 200
    except Exception as e:
        return jsonify({"code": 500, "msg": str(e), "data": None}), 500


def mcp_collect_douyin_account(homepage_url, max_videos=20):
    cookie_file = latest_douyin_cookie_file()
    return sync_one_douyin_benchmark(homepage_url, cookie_file, max_videos=max_videos)


def mcp_list_benchmark_videos(account_id, page=1, page_size=20):
    videos = benchmark_repository.list_videos(get_db_path(), account_id)
    page = max(1, int(page or 1))
    page_size = max(1, min(int(page_size or 20), 100))
    start = (page - 1) * page_size
    end = start + page_size
    return {
        "items": videos[start:end],
        "page": page,
        "page_size": page_size,
        "total": len(videos),
    }


def mcp_analyze_benchmark_video(video_id, force=False):
    video = benchmark_repository.get_video(get_db_path(), video_id)
    if not video:
        raise ValueError("作品不存在")
    existing = benchmark_repository.get_analysis(get_db_path(), video_id)
    if existing and not force and not is_douyin_analysis_stale(existing, video):
        return parse_douyin_video_analysis(existing)
    analysis = generate_douyin_video_analysis(video)
    save_douyin_video_analysis(video_id, analysis)
    return parse_douyin_video_analysis(benchmark_repository.get_analysis(get_db_path(), video_id))


def mcp_run_idea_radar(video_id, target_direction="AI 生产系统研究员", force=False, force_transcription=False):
    if not load_idea_radar_video(video_id):
        raise ValueError("作品不存在")
    return start_idea_radar_pipeline(
        video_id,
        target_direction,
        force=force,
        force_transcription=force_transcription,
    )


def mcp_generate_my_script(identity_profile, radar_result, benchmark_analysis=None):
    return generate_identity_script(identity_profile, radar_result, benchmark_analysis or {})


def mcp_review_published_content(work_id=None, limit=50):
    videos = list_own_douyin_videos(limit)
    if work_id is not None:
        videos = [video for video in videos if str(video.get("id")) == str(work_id)]
    return review_published_content(videos)


def mcp_tool_handlers():
    return create_tool_handlers({
        "collect_douyin_account": mcp_collect_douyin_account,
        "list_benchmark_videos": mcp_list_benchmark_videos,
        "analyze_benchmark_video": mcp_analyze_benchmark_video,
        "run_idea_radar": mcp_run_idea_radar,
        "generate_my_script": mcp_generate_my_script,
        "review_published_content": mcp_review_published_content,
    })


@app.route('/mcp/tools', methods=['GET'])
def list_mcp_tools():
    return jsonify({"code": 200, "msg": "success", "data": describe_mcp_tools()}), 200


@app.route('/mcp/invoke', methods=['POST'])
def invoke_mcp_tool():
    payload = request.get_json(silent=True) or {}
    tool_name = str(payload.get("name") or "").strip()
    arguments = payload.get("arguments") or {}
    if not tool_name:
        return jsonify({"code": 400, "msg": "请提供 MCP 工具名称", "data": None}), 400
    if not isinstance(arguments, dict):
        return jsonify({"code": 400, "msg": "MCP 工具参数必须是对象", "data": None}), 400
    handlers = mcp_tool_handlers()
    handler = handlers.get(tool_name)
    if not handler:
        return jsonify({"code": 404, "msg": "MCP 工具不存在", "data": None}), 404
    try:
        return jsonify({"code": 200, "msg": "success", "data": handler(arguments)}), 200
    except NotImplementedError as exc:
        return jsonify({"code": 501, "msg": str(exc), "data": None}), 501
    except KeyError as exc:
        return jsonify({"code": 400, "msg": f"缺少参数：{exc}", "data": None}), 400
    except Exception as exc:
        return jsonify({"code": 500, "msg": str(exc), "data": None}), 500


@app.route('/deleteFile', methods=['GET'])
def delete_file():
    file_id = request.args.get('id')

    if not file_id or not file_id.isdigit():
        return jsonify({
            "code": 400,
            "msg": "Invalid or missing file ID",
            "data": None
        }), 400

    try:
        record = material_repository.delete_file_record(get_db_path(), file_id)
        if not record:
            return jsonify({
                "code": 404,
                "msg": "File not found",
                "data": None
            }), 404

        return jsonify({
            "code": 200,
            "msg": "File deleted successfully",
            "data": {
                "id": record['id'],
                "filename": record['filename']
            }
        }), 200

    except Exception as e:
        return jsonify({
            "code": 500,
            "msg": str("delete failed!"),
            "data": None
        }), 500

@app.route('/deleteAccount', methods=['GET'])
def delete_account():
    account_id = int(request.args.get('id'))

    try:
        record = account_repository.delete_account(get_db_path(), account_id)
        if not record:
            return jsonify({
                "code": 404,
                "msg": "account not found",
                "data": None
            }), 404

        return jsonify({
            "code": 200,
            "msg": "account deleted successfully",
            "data": None
        }), 200

    except Exception as e:
        return jsonify({
            "code": 500,
            "msg": str("delete failed!"),
            "data": None
        }), 500


# SSE 登录接口
@app.route('/updateUserinfo', methods=['POST'])
def updateUserinfo():
    # 获取JSON数据
    data = request.get_json()

    # 从JSON数据中提取 type 和 userName
    user_id = data.get('id')
    type = data.get('type')
    userName = data.get('userName')
    try:
        account_repository.update_account_info(get_db_path(), user_id, type, userName)

        return jsonify({
            "code": 200,
            "msg": "account update successfully",
            "data": None
        }), 200

    except Exception as e:
        return jsonify({
            "code": 500,
            "msg": str("update failed!"),
            "data": None
        }), 500

if __name__ == '__main__':
    ensure_core_tables()
    ensure_default_agent_model()
    app.run(
        host=os.environ.get("SAU_BACKEND_HOST", "127.0.0.1"),
        port=int(os.environ.get("SAU_BACKEND_PORT", "5409")),
    )
